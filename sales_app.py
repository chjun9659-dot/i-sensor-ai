import os
import re
import io
from datetime import datetime, date, timedelta
from modules.ui_common import ui_card
import pandas as pd
import streamlit as st
from modules.ui_common import page_title 
from openpyxl import load_workbook
import calendar
import streamlit.components.v1 as components
from modules.ui_common import render_common_style, ui_card
def render_inspection_common_style():
    render_common_style()
from modules.page_schedule import schedule_page
from modules.page_inspection import inspection_page
from modules.page_dashboard import page_dashboard
from modules.page_maintenance import maintenance_page

st.set_page_config(page_title="윤우 영업 통합 시스템", layout="wide")
# =========================================================
# 기본 설정
# =========================================================
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

# =========================================================
# 사용자관리 구글시트
# =========================================================
USER_SHEET_URL = "https://docs.google.com/spreadsheets/d/1uUjrdRwTjdvKoED1dWKsikAOGWNxMR59Eht-SYdRX1w/edit?gid=0#gid=0"
# =========================================================
# 업무일정 구글시트
# =========================================================
# 업무일정 구글시트
WORK_SCHEDULE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1OfYbKceuwIcCqhGjDxT0nevQXSkNH7muwsKZOPiMHZs/edit?gid=0#gid=0"
NOTICE_SHEET_URL = "https://docs.google.com/spreadsheets/d/18D9HWhSzuf9Mfo4lPqD36nMsfYynjWUmyIuW4w7H0Sc/edit?gid=0#gid=0"
WORK_SCHEDULE_SHEET_NAME = "업무일정"
WORK_SCHEDULE_COLUMNS = ["날짜", "업무내용", "담당자", "상태", "메모"]

TODAY_TASK_SHEET_NAME = "오늘할일"
TODAY_TASK_COLUMNS = ["등록일시", "작성자", "사업", "할일", "상태"]

@st.cache_data(ttl=300)
def load_today_tasks():
    try:
        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", WORK_SCHEDULE_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(TODAY_TASK_SHEET_NAME)

        data = ws.get_all_records()
        df = pd.DataFrame(data)

        if df.empty:
            df = pd.DataFrame(columns=TODAY_TASK_COLUMNS)

        for col in TODAY_TASK_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        return df[TODAY_TASK_COLUMNS].copy()

    except Exception as e:
        st.warning(f"오늘할일 불러오기 실패: {e}")
        return pd.DataFrame(columns=TODAY_TASK_COLUMNS)

@st.cache_data(ttl=300)
def load_notice():
    try:
        client = get_gsheet_client()
        spreadsheet = client.open_by_url(NOTICE_SHEET_URL)
        ws = spreadsheet.worksheet("공지사항")

        data = ws.get_all_records()
        df = pd.DataFrame(data)

        return df

    except Exception as e:
        st.warning(f"공지사항 불러오기 실패: {e}")
        return pd.DataFrame(columns=["작성일", "내용", "작성자"])
    
def save_today_tasks(df):
    try:
        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", WORK_SCHEDULE_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        ws = spreadsheet.worksheet(TODAY_TASK_SHEET_NAME)

        save_df = df[TODAY_TASK_COLUMNS].fillna("")

        ws.clear()
        ws.update([TODAY_TASK_COLUMNS] + save_df.values.tolist())

        load_today_tasks.clear()

    except Exception as e:
        st.error(f"오늘할일 저장 실패: {e}") 

def load_users_from_gsheet():
    """
    사용자관리 시트 컬럼:
    아이디 / 비밀번호 / 권한 / 사용여부 / 이름 / 부서 / 직급 / 코드
    """
    try:
        if not USER_SHEET_URL or "여기에_" in USER_SHEET_URL:
            return {}

        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", USER_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(0)

        values = worksheet.get_all_values()

        if not values or len(values) < 2:
            st.error("사용자관리 시트에 데이터가 없습니다.")
            return {}

        headers = [str(x).strip() for x in values[0]]
        rows = values[1:]

        df = pd.DataFrame(rows, columns=headers).fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        required_cols = ["아이디", "비밀번호", "권한", "사용여부", "이름"]
        for col in required_cols:
            if col not in df.columns:
                st.error(f"사용자관리 시트에 '{col}' 컬럼이 없습니다.")
                return {}

        users = {}

        for _, row in df.iterrows():
            user_id = str(row.get("아이디", "")).strip()
            password = str(row.get("비밀번호", "")).strip()
            role = str(row.get("권한", "")).strip()   # 🔥 핵심 수정
            use_yn = str(row.get("사용여부", "")).strip().upper()
            name = str(row.get("이름", "")).strip()

            if not user_id:
                continue

            users[user_id] = {
                "pw": password,
                "role": role,   # 🔥 그대로 넣기
                "name": name,
                "use_yn": use_yn,
                "department": str(row.get("부서", "")).strip(),
                "position": str(row.get("직급", "")).strip(),
                "code": str(row.get("코드", "")).strip(),
            }

        return users

    except Exception as e:
        st.error(f"사용자관리 시트 불러오기 오류: {e}")
        return {}

# =========================================================
# 사업별 구글시트 링크
# =========================================================
BUSINESS_CONFIG = {
    "아이센서": {
        "sheets": {
            "영업현황": "https://docs.google.com/spreadsheets/d/1CWTvHC1r6i5wjcZoFJa5kAm-6T0SKao1EdjI8jwVQG8/edit?gid=167508641#gid=167508641",
            "가능단지": "https://docs.google.com/spreadsheets/d/1CWTvHC1r6i5wjcZoFJa5kAm-6T0SKao1EdjI8jwVQG8/edit?gid=1108943027#gid=1108943027",
            "입찰공고": "https://docs.google.com/spreadsheets/d/1CWTvHC1r6i5wjcZoFJa5kAm-6T0SKao1EdjI8jwVQG8/edit?gid=243967548#gid=243967548",
            "계약단지": "https://docs.google.com/spreadsheets/d/1CWTvHC1r6i5wjcZoFJa5kAm-6T0SKao1EdjI8jwVQG8/edit?gid=2071693391#gid=2071693391",
            "연차관리": "https://docs.google.com/spreadsheets/d/1n7AXfaCIljI8cBMSX7DrLjPVEQPTTCXL2-rxqcLzepE/edit?gid=0#gid=0",
            "시공일정": "https://docs.google.com/spreadsheets/d/1-6P8Orzas1U6W-Rmv7pcx-N0-fiPnH-Ah20jDEsRXgs/edit?gid=1427359982#gid=1427359982",
            "실사관리": "https://docs.google.com/spreadsheets/d/1rV9nZWGQDgUBgxldvojixj8Ys5DsefH9wa5-IWg_t34/edit?gid=859568227#gid=859568227",
        },
        "menus": [
            "대시보드",
            "데이터 가져오기",

            "영업현황",
            "가능단지",
            "입찰공고단지",
            "계약단지",
            "라우터 관리",
            "아이센서 유지보수관리",

            "연차 관리",
            "시공 일정",
            "실사 관리",

            "오늘 할 일",
            "일정 관리",
            "영업 알림",
        ],
    },
        "전기차 충전기": {
        "sheets": {
            "계약접수현황": "https://docs.google.com/spreadsheets/d/1Efuld8DUSHs8jR42R5XKIT3y-3P_pf--aP9ED_SYWhc/edit?gid=784400128#gid=784400128",
            "시공일정": "https://docs.google.com/spreadsheets/d/1-6P8Orzas1U6W-Rmv7pcx-N0-fiPnH-Ah20jDEsRXgs/edit?gid=1427359982#gid=1427359982",
            "실사관리": "https://docs.google.com/spreadsheets/d/1rV9nZWGQDgUBgxldvojixj8Ys5DsefH9wa5-IWg_t34/edit?gid=859568227#gid=859568227",
        },
        "menus": [
            "대시보드",
            "데이터 가져오기",
            "계약접수현황",
            "시공 일정",
            "실사 관리",
            "오늘 할 일",
            "일정 관리",
            "영업 알림",
        ],
    },
} 

# =========================================================
# 세션 상태
# =========================================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "username" not in st.session_state:
    st.session_state.username = ""
if "role" not in st.session_state:
    st.session_state.role = ""
if "display_name" not in st.session_state:
    st.session_state.display_name = ""
if "business" not in st.session_state:
    st.session_state.business = "아이센서"

if "department" not in st.session_state:
    st.session_state.department = ""

if "position" not in st.session_state:
    st.session_state.position = ""

if "user_code" not in st.session_state:
    st.session_state.user_code = ""

if "inspection_form_version" not in st.session_state:
    st.session_state.inspection_form_version = 0

if "inspection_edit_mode" not in st.session_state:
    st.session_state.inspection_edit_mode = False

if "inspection_edit_target" not in st.session_state:
    st.session_state.inspection_edit_target = None    


# =========================================================
# 공통 유틸
# =========================================================
def clean_text(value):
    if pd.isna(value):
        return ""
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]", "", value)
    return value


def make_unique_columns(cols):
    seen = {}
    new_cols = []
    for c in cols:
        base = str(c).strip()
        if base == "":
            base = "빈컬럼"
        if base not in seen:
            seen[base] = 0
            new_cols.append(base)
        else:
            seen[base] += 1
            new_cols.append(f"{base}_{seen[base]}")
    return new_cols


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    cols = []
    for c in df.columns:
        c = str(c).replace("\n", " ").strip()
        if c.startswith("Unnamed"):
            c = ""
        cols.append(c)
    df.columns = make_unique_columns(cols)
    return df


def remove_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    keep_cols = []
    for col in df.columns:
        if str(col).strip() in ["", "빈컬럼"]:
            continue
        col_data = df[col]
        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]
        if not col_data.isna().all():
            keep_cols.append(col)
    return df[keep_cols].copy()


def normalize_text(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def format_date_value(x):
    if pd.isna(x) or x == "":
        return ""
    try:
        if isinstance(x, pd.Timestamp):
            return x.strftime("%Y-%m-%d")
        if isinstance(x, datetime):
            return x.strftime("%Y-%m-%d")
        s = str(x).strip()
        s = s.replace(".", "-").replace("/", "-")
        return s
    except Exception:
        return str(x)


def preprocess_df(df: pd.DataFrame) -> pd.DataFrame:
    df = clean_columns(df)
    df = df.fillna("")

    for col in df.columns:
        col_data = df[col]
        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]

        if any(key in str(col) for key in ["날짜", "일자", "마감", "공고", "접수", "실사", "발행", "입금", "계약"]):
            df[col] = col_data.apply(format_date_value)
        else:
            df[col] = col_data.apply(lambda x: x if isinstance(x, (int, float)) else normalize_text(x))
    return df


def business_prefix():
    return "sensor" if st.session_state.business == "아이센서" else "ev"


def local_file_path(sheet_key: str) -> str:
    prefix = business_prefix()
    safe_name = re.sub(r"[^가-힣a-zA-Z0-9_]", "_", sheet_key)
    return os.path.join(DATA_DIR, f"{prefix}_{safe_name}.csv")


def save_df(sheet_key: str, df: pd.DataFrame):
    path = local_file_path(sheet_key)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def load_local_df(sheet_key: str) -> pd.DataFrame:
    path = local_file_path(sheet_key)
    if os.path.exists(path):
        try:
            return pd.read_csv(path, encoding="utf-8-sig").fillna("")
        except Exception:
            return pd.read_csv(path).fillna("")
    return pd.DataFrame()


def get_current_sheet_urls():
    return BUSINESS_CONFIG[st.session_state.business]["sheets"]


def log_event(action, detail=""):
    print(f"[{datetime.now()}] {action} | {detail}")


def backup_before_save(df, name):
    os.makedirs("backup", exist_ok=True)
    path = f"backup/{name}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    df.to_csv(path, index=False, encoding="utf-8-sig")


def gsheet_read(sheet_name, url):
    try:
        df = load_google_sheet_data(st.session_state.business, sheet_name, url)
        log_event("READ", f"{sheet_name} 성공")
        return df
    except Exception as e:
        log_event("READ_FAIL", f"{sheet_name} / {e}")
        return pd.DataFrame()


def gsheet_write(sheet_name, df, url=None):
    try:
        # 1) 저장 전 백업
        backup_before_save(df, sheet_name)

        # 2) 기존 로컬 저장
        save_df(sheet_name, df)

        # 3) 캐시 초기화
        try:
            load_google_sheet_data.clear()
        except:
            pass

        try:
            load_today_tasks.clear()
        except:
            pass

        try:
            load_notice.clear()
        except:
            pass

        log_event("WRITE", f"{sheet_name} 저장완료 / 캐시초기화 완료")

    except Exception as e:
        log_event("WRITE_FAIL", f"{sheet_name} / {e}")

# =========================================================
# 구글 시트 연동
# =========================================================
import gspread
from google.oauth2.service_account import Credentials

def get_gsheet_client():
    import os
    import gspread
    from google.oauth2.service_account import Credentials
    import streamlit as st

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    base_dir = os.path.dirname(os.path.abspath(__file__))
    key_path = os.path.join(base_dir, "service_account.json")

    try:
        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(
            dict(creds_dict),
            scopes=scopes
        )

        client = gspread.authorize(creds)

        st.session_state["google_auth_debug"] = {
            "mode": "streamlit_secrets",
            "client_email": creds.service_account_email,
        }

        return client

    except Exception as secrets_error:

        if os.path.exists(key_path):

            creds = Credentials.from_service_account_file(
                key_path,
                scopes=scopes
            )

            client = gspread.authorize(creds)

            st.session_state["google_auth_debug"] = {
                "mode": "local_json",
                "key_path": key_path,
                "client_email": creds.service_account_email,
                "secrets_error": str(secrets_error),
            }

            return client

        raise RuntimeError(
            f"구글 인증 실패 / secrets 오류: {secrets_error}"
        )
    
def get_drive_service():
    from googleapiclient.discovery import build
    from google.oauth2.service_account import Credentials
    import os
    import streamlit as st

    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
    ]

    base_dir = os.path.dirname(os.path.abspath(__file__))
    key_path = os.path.join(base_dir, "service_account.json")

    try:
        creds_dict = st.secrets["gcp_service_account"]
        creds = Credentials.from_service_account_info(dict(creds_dict), scopes=scopes)
    except Exception:
        creds = Credentials.from_service_account_file(key_path, scopes=scopes)

    return build("drive", "v3", credentials=creds)

def upload_file_to_drive(file, folder_id=None):
    from googleapiclient.http import MediaIoBaseUpload
    import streamlit as st

    try:
        if file is None:
            return ""

        service = get_drive_service()

        file_metadata = {
            "name": file.name
        }

        if folder_id:
            file_metadata["parents"] = [folder_id]

        media = MediaIoBaseUpload(
            file,
            mimetype=file.type,
            resumable=True
        )

        uploaded = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, webViewLink"
        ).execute()

        return uploaded.get("webViewLink", "")

    except Exception as e:
        st.error(f"첨부파일 업로드 실패: {e}")
        return ""

    except Exception as secrets_error:
        # 2) 로컬 fallback: service_account.json 이 실제 있을 때만 사용
        if os.path.exists(key_path):
            creds = Credentials.from_service_account_file(key_path, scopes=scopes)
            client = gspread.authorize(creds)

            st.session_state["google_auth_debug"] = {
                "mode": "local_json",
                "key_path": key_path,
                "client_email": creds.service_account_email,
                "secrets_error": str(secrets_error),
            }
            return client

        # 3) 둘 다 없으면 실제 원인을 보여줌
        raise RuntimeError(
            f"구글 인증 실패 / secrets 오류: {secrets_error} / "
            f"로컬 JSON 없음: {key_path}"
        )

    raise FileNotFoundError(
        "구글 인증 정보를 찾지 못했습니다. "
        "배포환경은 st.secrets['gcp_service_account'], "
        "로컬환경은 service_account.json 이 필요합니다."
    )

    
def append_to_gsheet(sheet_url, row_data, worksheet_index=0):
    try:
        client = get_gsheet_client()

        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", sheet_url).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(worksheet_index)
        worksheet.append_row(row_data)

        return True
    except Exception as e:
        st.error(f"구글 저장 오류: {e}")
        return False

def update_billing_status_in_gsheet(sheet_url, 기준월, 단지명, 담당자, 청구금액, worksheet_index=0):
    try:
        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", sheet_url).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(worksheet_index)

        values = worksheet.get_all_values()
        if not values or len(values) < 2:
            st.session_state["google_update_msg"] = "구글 시트에 데이터가 없습니다."
            return False

        target_ym = str(기준월).strip()
        target_name = str(단지명).strip()
        target_manager = str(담당자).strip()

        for i, row in enumerate(values[1:], start=2):
            row_기준월 = str(row[0]).strip() if len(row) > 0 else ""
            row_단지명 = str(row[1]).strip() if len(row) > 1 else ""
            row_담당자 = str(row[2]).strip() if len(row) > 2 else ""

            # 청구금액은 비교하지 않음
            if (
                row_기준월 == target_ym and
                row_단지명 == target_name and
                row_담당자 == target_manager
            ):
                worksheet.update_cell(i, 5, "입금")   # E열
                st.session_state["google_update_msg"] = (
                    f"구글 시트 업데이트 완료: {target_name} / 행={i}"
                )
                return True

        st.session_state["google_update_msg"] = (
            f"일치 행을 찾지 못했습니다: 기준월={target_ym}, 단지명={target_name}, 담당자={target_manager}"
        )
        return False

    except Exception as e:
        st.session_state["google_update_msg"] = f"구글 수금관리 업데이트 오류: {type(e).__name__} / {e}"
        return False

def convert_google_sheet_url_to_csv(url: str) -> str:
    sheet_match = re.search(r"/d/([a-zA-Z0-9-_]+)", url)
    gid_match = re.search(r"gid=(\d+)", url)

    if not sheet_match:
        raise ValueError("구글 스프레드시트 문서 ID를 찾을 수 없습니다.")

    sheet_id = sheet_match.group(1)
    gid = gid_match.group(1) if gid_match else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv&gid={gid}"


def detect_header_row(raw: pd.DataFrame) -> int:
    header_keywords = [
        "관리코드", "관리번호", "아파트명", "아파트 명", "단지명", "현장명",
        "연락처", "지역", "주소", "상품", "판매형태", "운영계약", "결제방식",
        "영업담당", "담당자", "실사담당", "수량", "판매가격", "계약날짜",
        "행위신고", "시공요청서", "시공", "시공여부", "세금계산서 발행",
        "세금계산서발행", "입금일", "영업수수료", "운영사", "구분",
        "계약서 유무", "서류 풀 세팅 완료 표시", "추가요금", "설치업체", "설치유무"
    ]

    if raw.empty:
        return 0

    scan_rows = min(len(raw), 10)
    best_score = -1
    best_idx = 0

    for i in range(scan_rows):
        row_values = raw.iloc[i].astype(str).str.strip().tolist()
        score = sum(1 for v in row_values if v in header_keywords)
        if score > best_score:
            best_score = score
            best_idx = i

    return best_idx


def force_fix_quantity_column(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    df = df.copy()

    def is_currency_only_name(name: str) -> bool:
        return str(name).strip() in ["₩", "￦"]

    def is_effective_blank(series: pd.Series) -> bool:
        s = series.astype(str).str.strip()
        return (s == "").all()

    def rename_sensor_contract_columns(df_inner: pd.DataFrame) -> pd.DataFrame:
        cols = list(df_inner.columns)

        if "영업담당" not in cols:
            return df_inner

        manager_idx = cols.index("영업담당")
        right_cols = cols[manager_idx + 1:]

        usable_cols = []
        for c in right_cols:
            col_series = df_inner[c]
            if isinstance(col_series, pd.DataFrame):
                col_series = col_series.iloc[:, 0]

            col_name = str(c).strip()

            if is_currency_only_name(col_name):
                continue

            if is_effective_blank(col_series):
                continue

            usable_cols.append(c)

        expected = [
            "수량",
            "판매가격",
            "계약날짜",
            "행위신고",
            "시공요청서",
            "시공",
            "시공여부",
            "세금계산서 발행",
            "입금일",
            "영업수수료",
        ]

        rename_map = {}
        for old, new in zip(usable_cols, expected):
            rename_map[old] = new

        if rename_map:
            df_inner = df_inner.rename(columns=rename_map)
            df_inner.columns = make_unique_columns(df_inner.columns)

        drop_cols = [c for c in df_inner.columns if is_currency_only_name(c)]
        if drop_cols:
            df_inner = df_inner.drop(columns=drop_cols, errors="ignore")  
            
        # -----------------------------------------
        # 라우터 컬럼명 보정
        # -----------------------------------------
        cols_after = list(df_inner.columns)

        if "라우터청구대상" in cols_after:
            billing_idx = cols_after.index("라우터청구대상")

            if billing_idx + 1 < len(cols_after):
                next_col = cols_after[billing_idx + 1]
                if str(next_col).startswith("빈컬럼"):
                    df_inner = df_inner.rename(columns={next_col: "라우터월비용"})

            cols_after = list(df_inner.columns)
            if billing_idx + 2 < len(cols_after):
                next_col = cols_after[billing_idx + 2]
                if str(next_col).startswith("빈컬럼"):
                    df_inner = df_inner.rename(columns={next_col: "라우터청구시작월"})

            cols_after = list(df_inner.columns)
            if billing_idx + 3 < len(cols_after):
                next_col = cols_after[billing_idx + 3]
                if str(next_col).startswith("빈컬럼"):
                    df_inner = df_inner.rename(columns={next_col: "라우터청구종료월"})

        df_inner.columns = make_unique_columns(df_inner.columns)
        return df_inner

    def rename_ev_contract_columns(df_inner: pd.DataFrame) -> pd.DataFrame:
        cols = list(df_inner.columns)

        if "주소" not in cols:
            return df_inner

        addr_idx = cols.index("주소")
        rename_map = {}

        # 주소 다음 컬럼들 강제 보정
        if addr_idx + 1 < len(cols):
            rename_map[cols[addr_idx + 1]] = "설치대수"
        if addr_idx + 2 < len(cols):
            rename_map[cols[addr_idx + 2]] = "주차면"
        if addr_idx + 3 < len(cols):
            rename_map[cols[addr_idx + 3]] = "계약서 유무"
        if addr_idx + 4 < len(cols):
            rename_map[cols[addr_idx + 4]] = "서류 풀 세팅 완료 표시"
        if addr_idx + 5 < len(cols):
            rename_map[cols[addr_idx + 5]] = "추가요금"
        if addr_idx + 6 < len(cols):
            rename_map[cols[addr_idx + 6]] = "설치업체"
        if addr_idx + 7 < len(cols):
            rename_map[cols[addr_idx + 7]] = "설치유무"
        if addr_idx + 8 < len(cols):
            rename_map[cols[addr_idx + 8]] = "계약날짜"
        if addr_idx + 9 < len(cols):
            rename_map[cols[addr_idx + 9]] = "원본 등기 발송"
        if addr_idx + 10 < len(cols):
            rename_map[cols[addr_idx + 10]] = "계약기간"
        if addr_idx + 11 < len(cols):
            rename_map[cols[addr_idx + 11]] = "운영사 접수"
        if addr_idx + 12 < len(cols):
            rename_map[cols[addr_idx + 12]] = "기타"
        if addr_idx + 13 < len(cols):
            rename_map[cols[addr_idx + 13]] = "전력인입"
        if addr_idx + 14 < len(cols):
            rename_map[cols[addr_idx + 14]] = "특이사항"

        df_inner = df_inner.rename(columns=rename_map)
        df_inner.columns = make_unique_columns(df_inner.columns)

        # 숫자 컬럼 정리
        for num_col in ["설치대수", "주차면"]:
            if num_col in df_inner.columns:
                num = pd.to_numeric(
                    df_inner[num_col].astype(str).str.replace(",", "", regex=False).str.strip(),
                    errors="coerce"
                )
                if num.notna().sum() > 0:
                    df_inner[num_col] = num.apply(
                        lambda x: int(x) if pd.notna(x) and float(x).is_integer() else x
                    )

        return df_inner

    # 아이센서 계약단지 전용
    if st.session_state.business == "아이센서" and sheet_name == "계약단지":
        df = rename_sensor_contract_columns(df)

        if "수량" in df.columns:
            qty = pd.to_numeric(
                df["수량"].astype(str).str.replace(",", "", regex=False).str.strip(),
                errors="coerce"
            )
            if qty.notna().sum() > 0:
                df["수량"] = qty.apply(
                    lambda x: int(x) if pd.notna(x) and float(x).is_integer() else x
                )

        if "판매가격" in df.columns:
            price = pd.to_numeric(
                df["판매가격"].astype(str)
                .str.replace(",", "", regex=False)
                .str.replace("₩", "", regex=False)
                .str.replace("￦", "", regex=False)
                .str.strip(),
                errors="coerce"
            )
            if price.notna().sum() > 0:
                df["판매가격"] = price.apply(
                    lambda x: int(x) if pd.notna(x) and float(x).is_integer() else x
                )

        if "영업수수료" in df.columns:
            fee = pd.to_numeric(
                df["영업수수료"].astype(str)
                .str.replace(",", "", regex=False)
                .str.replace("₩", "", regex=False)
                .str.replace("￦", "", regex=False)
                .str.strip(),
                errors="coerce"
            )
            if fee.notna().sum() > 0:
                df["영업수수료"] = fee.apply(
                    lambda x: int(x) if pd.notna(x) and float(x).is_integer() else x
                )

        return df

    # 전기차 계약접수현황 전용
    if st.session_state.business == "전기차 충전기" and sheet_name == "계약접수현황":
        df = rename_ev_contract_columns(df)
        return df

    # 기타 공통 처리
    if "수량" in df.columns:
        qty = pd.to_numeric(
            df["수량"].astype(str).str.replace(",", "", regex=False).str.strip(),
            errors="coerce"
        )
        if qty.notna().sum() > 0:
            df["수량"] = qty.apply(
                lambda x: int(x) if pd.notna(x) and float(x).is_integer() else x
            )

    return df


@st.cache_data(ttl=300)
def load_google_sheet_data(business_name: str, sheet_name: str, url: str) -> pd.DataFrame:
    if not url or "여기에_" in url:
        return pd.DataFrame()

    if sheet_name == "연차관리":
        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", url).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(0)

        values = worksheet.get_all_values()

        if not values or len(values) < 2:
            return pd.DataFrame()

        headers = [str(x).strip() for x in values[0]]
        rows = values[1:]

        df = pd.DataFrame(rows, columns=headers).fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        return df

    csv_url = convert_google_sheet_url_to_csv(url)
    raw = pd.read_csv(csv_url, header=None, dtype=str).fillna("")

    if raw.empty:
        return pd.DataFrame()

    header_row_idx = detect_header_row(raw)

    headers = raw.iloc[header_row_idx].astype(str).str.strip().tolist()
    data = raw.iloc[header_row_idx + 1:].reset_index(drop=True).copy()
    data.columns = headers

    df = data.copy()
    df = preprocess_df(df)
    df = force_fix_quantity_column(df, sheet_name)

    return df   # 🔥 이거 추가

def load_df(sheet_key: str) -> pd.DataFrame:
    sheet_urls = get_current_sheet_urls()

    if sheet_key in sheet_urls:
        try:
            df = gsheet_read(sheet_key, sheet_urls[sheet_key])
            if not df.empty:
                gsheet_write(sheet_key, df)
                return df
        except Exception as e:
            st.warning(f"{sheet_key} 구글시트 불러오기 실패, 로컬 백업 사용: {e}")

    return load_local_df(sheet_key)


# =========================================================
# 공통 보조 파일 (사업 공통)
# =========================================================
COMMON_FILE_MAP = {
    "할일": os.path.join(DATA_DIR, "common_tasks.csv"),
    "일정": os.path.join(DATA_DIR, "common_schedule.csv"),
    "세금알림": os.path.join(DATA_DIR, "common_tax_alerts.csv"),
    "입대의알림": os.path.join(DATA_DIR, "common_meeting_alerts.csv"),
}


def save_common_df(key: str, df: pd.DataFrame):
    df.to_csv(COMMON_FILE_MAP[key], index=False, encoding="utf-8-sig")


def load_common_df(key: str) -> pd.DataFrame:
    path = COMMON_FILE_MAP[key]
    if os.path.exists(path):
        try:
            return pd.read_csv(path, encoding="utf-8-sig").fillna("")
        except Exception:
            return pd.read_csv(path).fillna("")
    return pd.DataFrame()


def ensure_common_file(key: str, columns: list[str]):
    if not os.path.exists(COMMON_FILE_MAP[key]):
        save_common_df(key, pd.DataFrame(columns=columns))


def init_files():
    ensure_common_file("할일", ["등록일시", "작성자", "사업", "할일"])
    ensure_common_file("일정", ["등록일시", "작성자", "사업", "일정명", "날짜"])
    ensure_common_file("세금알림", ["등록일시", "작성자", "사업", "단지명", "예정일", "상태", "비고"])
    ensure_common_file("입대의알림", ["등록일시", "작성자", "사업", "단지명", "입대의일자", "상태", "비고"])
    ensure_money_files()

def load_tasks_df():
    df = load_today_tasks()
    for col in ["등록일시", "작성자", "사업", "할일", "상태"]:
        if col not in df.columns:
            df[col] = ""
    return df[["등록일시", "작성자", "사업", "할일", "상태"]].copy()


def save_tasks_df(df: pd.DataFrame):
    save_today_tasks(df)

def load_schedule_df():
    df = load_work_schedule_data()

    # 기존 업무일정 화면 컬럼 구조 유지
    rename_map = {
        "업무내용": "일정명",
        "담당자": "작성자",
        "메모": "사업",
    }

    df = df.rename(columns=rename_map)

    for col in ["등록일시", "작성자", "사업", "일정명", "날짜"]:
        if col not in df.columns:
            df[col] = ""

    return df[["등록일시", "작성자", "사업", "일정명", "날짜"]].copy()

def save_schedule_df(df: pd.DataFrame):
    save_df = df.copy()

    save_df["업무내용"] = save_df.get("일정명", "")
    save_df["담당자"] = save_df.get("작성자", "")
    save_df["상태"] = save_df.get("상태", "")
    save_df["메모"] = save_df.get("사업", "")
    save_df["날짜"] = save_df.get("날짜", "")  # ✅ 추가

    save_work_schedule_data(
        save_df[["날짜", "업무내용", "담당자", "상태", "메모"]].copy()
    )


def load_tax_alert_df():
    df = load_common_df("세금알림")
    for col in ["등록일시", "작성자", "사업", "단지명", "예정일", "상태", "비고"]:
        if col not in df.columns:
            df[col] = ""
    return df[["등록일시", "작성자", "사업", "단지명", "예정일", "상태", "비고"]].copy()


def save_tax_alert_df(df: pd.DataFrame):
    save_common_df("세금알림", df[["등록일시", "작성자", "사업", "단지명", "예정일", "상태", "비고"]].copy())


def load_meeting_alert_df():
    df = load_common_df("입대의알림")
    for col in ["등록일시", "작성자", "사업", "단지명", "입대의일자", "상태", "비고"]:
        if col not in df.columns:
            df[col] = ""
    return df[["등록일시", "작성자", "사업", "단지명", "입대의일자", "상태", "비고"]].copy()


def save_meeting_alert_df(df: pd.DataFrame):
    save_common_df("입대의알림", df[["등록일시", "작성자", "사업", "단지명", "입대의일자", "상태", "비고"]].copy())

# =========================================================
# 수금관리 / 미입금관리 / 담당자별현황
# =========================================================
MONEY_FILE_MAP = {
    "수금관리": os.path.join(DATA_DIR, "sensor_billing_master.csv"),
    "미입금관리": os.path.join(DATA_DIR, "sensor_unpaid.csv"),
    "담당자별현황": os.path.join(DATA_DIR, "sensor_manager_summary.csv"),
}

# 실사용 구글 수금관리 시트 URL
BILLING_SHEET_URL = "https://docs.google.com/spreadsheets/d/1QDf1No9Nz5CVu3BGxLr7omWNSLi8Pth7FKTBChh6ds0/edit?gid=970462186#gid=970462186"


def save_money_df(key: str, df: pd.DataFrame):
    path = MONEY_FILE_MAP[key]
    df.to_csv(path, index=False, encoding="utf-8-sig")


def load_money_df(key: str) -> pd.DataFrame:
    path = MONEY_FILE_MAP[key]
    if os.path.exists(path):
        try:
            return pd.read_csv(path, encoding="utf-8-sig").fillna("")
        except Exception:
            return pd.read_csv(path).fillna("")
    return pd.DataFrame()


def ensure_money_files():
    if not os.path.exists(MONEY_FILE_MAP["수금관리"]):
        save_money_df(
            "수금관리",
            pd.DataFrame(columns=[
                "기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"
            ])
        )

    if not os.path.exists(MONEY_FILE_MAP["미입금관리"]):
        save_money_df(
            "미입금관리",
            pd.DataFrame(columns=[
                "기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"
            ])
        )

    if not os.path.exists(MONEY_FILE_MAP["담당자별현황"]):
        save_money_df(
            "담당자별현황",
            pd.DataFrame(columns=[
                "담당자", "미수금합계"
            ])
        )


def normalize_payment_status(value):
    s = str(value).strip()
    if s in ["입금", "입금완료", "완료", "수금완료", "Y", "y", "yes", "Yes"]:
        return "입금"
    return ""


def safe_int(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return default
    return int(num)


def load_billing_from_gsheet(sheet_url, worksheet_index=0) -> pd.DataFrame:
    try:
        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", sheet_url).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(worksheet_index)

        values = worksheet.get_all_values()
        if not values or len(values) < 2:
            return pd.DataFrame(columns=["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"])

        headers = [str(x).strip() for x in values[0]]
        rows = values[1:]
        df = pd.DataFrame(rows, columns=headers)

        for col in ["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"]:
            if col not in df.columns:
                df[col] = ""

        df = df[["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"]].copy()

        df["기준월"] = df["기준월"].astype(str).str.strip()
        df["단지명"] = df["단지명"].astype(str).str.strip()
        df["담당자"] = df["담당자"].astype(str).str.strip()
        df["입금여부"] = df["입금여부"].astype(str).apply(normalize_payment_status)

        df["청구금액"] = pd.to_numeric(df["청구금액"], errors="coerce").fillna(0)
        df["미수금"] = pd.to_numeric(df["미수금"], errors="coerce").fillna(0)

        df.loc[(df["청구금액"] <= 0) & (df["미수금"] > 0), "청구금액"] = df["미수금"]
        df.loc[(df["미수금"] <= 0) & (df["청구금액"] > 0) & (df["입금여부"] != "입금"), "미수금"] = df["청구금액"]
        df.loc[df["입금여부"] == "입금", "미수금"] = 0

        df["청구금액"] = df["청구금액"].astype(int)
        df["미수금"] = df["미수금"].astype(int)

        return df

    except Exception as e:
        st.session_state["google_update_msg"] = f"수금관리 구글시트 불러오기 오류: {type(e).__name__} / {e}"
        return pd.DataFrame(columns=["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"])

def build_billing_rows_from_router_claim_df(claim_df: pd.DataFrame) -> pd.DataFrame:
    if claim_df is None or claim_df.empty:
        return pd.DataFrame(columns=[
            "기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"
        ])

    df = claim_df.copy()

    if "청구년월" not in df.columns:
        df["청구년월"] = ""

    if "단지명" not in df.columns:
        df["단지명"] = ""

    if "담당자" not in df.columns:
        df["담당자"] = ""

    if "라우터월비용" not in df.columns:
        df["라우터월비용"] = 0

    df["청구금액"] = pd.to_numeric(df["라우터월비용"], errors="coerce").fillna(0).astype(int)
    df["입금여부"] = ""
    df["미수금"] = df["청구금액"]

    out = df[["청구년월", "단지명", "담당자", "청구금액", "입금여부", "미수금"]].copy()
    out = out.rename(columns={"청구년월": "기준월"})

    return out


def add_monthly_billing_data(claim_df: pd.DataFrame):
    """
    계약단지에서 추출한 이번달 청구대상을
    구글 실사용시트(BILLING_SHEET_URL)에 직접 추가한다.
    이미 같은 기준월 + 단지명 + 담당자 행이 있으면 중복 생성하지 않는다.
    """
    try:
        if claim_df is None or claim_df.empty:
            return {
                "added_count": 0,
                "duplicate_count": 0,
                "total_count": 0
            }

        client = get_gsheet_client()
        sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", BILLING_SHEET_URL).group(1)
        spreadsheet = client.open_by_key(sheet_id)
        worksheet = spreadsheet.get_worksheet(0)

        values = worksheet.get_all_values()
        if not values:
            # 헤더가 아예 없으면 생성
            worksheet.append_row(["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"])
            values = [["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"]]

        headers = [str(x).strip() for x in values[0]]
        rows = values[1:] if len(values) > 1 else []

        # 기존 행의 중복키 수집
        existing_row_map = {}

        for idx, row in enumerate(rows, start=2):
            row_기준월 = str(row[0]).strip() if len(row) > 0 else ""
            row_단지명 = str(row[1]).strip() if len(row) > 1 else ""
            row_담당자 = str(row[2]).strip() if len(row) > 2 else ""
            row_청구금액 = str(row[3]).strip() if len(row) > 3 else ""

            key = f"{row_기준월}||{row_단지명}||{row_담당자}"

            existing_row_map[key] = {
                "row_index": idx,
                "amount": row_청구금액
            }

        add_rows = []
        duplicate_count = 0

        work_df = claim_df.copy()

        if "청구년월" not in work_df.columns:
            work_df["청구년월"] = ""

        if "단지명" not in work_df.columns:
            work_df["단지명"] = ""

        if "담당자" not in work_df.columns:
            work_df["담당자"] = ""

        if "라우터월비용" not in work_df.columns:
            work_df["라우터월비용"] = 0

        for _, r in work_df.iterrows():
            기준월 = str(r.get("청구년월", "")).strip()
            단지명 = str(r.get("단지명", "")).strip()
            담당자 = str(r.get("담당자", "")).strip()
            청구금액 = int(pd.to_numeric(r.get("라우터월비용", 0), errors="coerce") or 0)

            if not 기준월 or not 단지명:
                continue

            if 청구금액 <= 0:
                continue

            key = f"{기준월}||{단지명}||{담당자}"

            if key in existing_row_map:
                row_info = existing_row_map[key]
                old_amount = row_info["amount"]
                row_index = row_info["row_index"]

                # 🔥 핵심: 금액 없으면 업데이트
                if old_amount in ["", "0"]:
                    worksheet.update_cell(row_index, 4, 청구금액)  # D열
                    worksheet.update_cell(row_index, 5, "")        # E열 초기화
                else:
                    duplicate_count += 1

                continue

            add_rows.append([
                기준월,
                단지명,
                담당자,
                청구금액,
                "",         # 입금여부
                청구금액    # 미수금
            ])
            existing_row_map[key] = {
                "row_index": None,
                "amount": str(청구금액)
            }

        if add_rows:
            worksheet.append_rows(add_rows, value_input_option="USER_ENTERED")

        total_count = len(rows) + len(add_rows)

        st.session_state["google_update_msg"] = (
            f"자동청구 생성 완료 / 신규 {len(add_rows)}건 / 중복 제외 {duplicate_count}건"
        )

        return {
            "added_count": len(add_rows),
            "duplicate_count": duplicate_count,
            "total_count": total_count
        }

    except Exception as e:
        st.session_state["google_update_msg"] = f"자동청구 생성 오류: {type(e).__name__} / {e}"
        return {
            "added_count": 0,
            "duplicate_count": 0,
            "total_count": 0
        }


def rebuild_billing_views():
    """
    기존 로컬 기반 함수는 더 이상 핵심 로직에서 사용하지 않음.
    남겨두되 동작은 최소화.
    """
    return


def load_billing_dashboard_data():
    """
    앱 수금관리 화면은 이제 구글 실사용시트를 원본으로 사용
    """
    billing_df = load_billing_from_gsheet(BILLING_SHEET_URL)

    if billing_df.empty:
        unpaid_df = pd.DataFrame(columns=["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"])
        manager_df = pd.DataFrame(columns=["담당자", "미수금합계"])
        return billing_df, unpaid_df, manager_df

    unpaid_df = billing_df[billing_df["미수금"] > 0].copy()
    unpaid_df = unpaid_df[["기준월", "단지명", "담당자", "청구금액", "입금여부", "미수금"]].copy()

    if unpaid_df.empty:
        manager_df = pd.DataFrame(columns=["담당자", "미수금합계"])
    else:
        manager_df = (
            unpaid_df.groupby("담당자", dropna=False)["미수금"]
            .sum()
            .reset_index()
            .rename(columns={"미수금": "미수금합계"})
            .sort_values("미수금합계", ascending=False)
        )

    return billing_df, unpaid_df, manager_df


def mark_billing_paid(기준월, 단지명, 담당자, 청구금액):
    """
    입금 처리는 구글 실사용시트만 수정
    """
    amount_num = pd.to_numeric(청구금액, errors="coerce")
    if pd.isna(amount_num):
        st.session_state["google_update_msg"] = f"청구금액이 비어 있거나 숫자가 아닙니다: {청구금액}"
        return False

    return update_billing_status_in_gsheet(
        BILLING_SHEET_URL,
        str(기준월).strip(),
        str(단지명).strip(),
        str(담당자).strip(),
        int(amount_num)
    )

# =========================================================
# 로그인
# =========================================================
def login():
    st.title("🔐 윤우 영업 통합 시스템 로그인")

    user_id = st.text_input("아이디")
    password = st.text_input("비밀번호", type="password")

    if st.button("로그인"):
        users = load_users_from_gsheet()

        if not users:
            st.error("사용자 정보를 불러오지 못했습니다. 사용자관리 시트를 확인하세요.")
            return

        if user_id not in users:
            st.error("아이디 또는 비밀번호가 맞지 않습니다.")
            return

        user = users[user_id]

        if str(user.get("use_yn", "")).upper() != "Y":
            st.error("사용 중지된 계정입니다. 관리자에게 문의하세요.")
            return

        if user.get("pw") != password:
            st.error("아이디 또는 비밀번호가 맞지 않습니다.")
            return

        st.session_state.logged_in = True
        st.session_state.username = user_id
        st.session_state.role = user.get("role", "user")
        st.session_state.display_name = user.get("name", user_id)
        st.session_state.department = user.get("department", "")
        st.session_state.position = user.get("position", "")
        st.session_state.user_code = user.get("code", "")

        st.rerun()


# =========================================================
# 컬럼 도우미
# =========================================================
def get_best_column(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def get_site_column(df):
    return get_best_column(df, ["아파트 명", "아파트명", "단지명", "현장명", "주소"])


def get_manager_column(df):
    return get_best_column(df, ["담당자", "영업담당", "실사담당", "작성자"])


def get_code_column(df):
    return get_best_column(df, ["관리코드", "관리 코드"])


# =========================================================
# 권한 필터
# =========================================================
def apply_role_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or is_admin():
        return df

    manager_col = get_manager_column(df)
    if not manager_col:
        return df

    user_name = current_user_name().strip()
    return df[df[manager_col].astype(str).str.strip() == user_name].copy()


def apply_author_filter(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    # ✅ 업무관리는 사업 구분 없이 작성자 기준만 적용
    if is_admin():
        return df

    if "작성자" not in df.columns:
        return df

    user_name = current_user_name().strip()

    return df[
        df["작성자"].astype(str).str.strip() == user_name
    ].copy()


# =========================================================
# 화면 공통
# =========================================================
def filtered_df(df: pd.DataFrame, filters: dict) -> pd.DataFrame:
    result = df.copy()
    for col, value in filters.items():
        if col in result.columns and value and value != "전체":
            result = result[result[col].astype(str) == str(value)]
    return result


def keyword_filter(df: pd.DataFrame, keyword: str) -> pd.DataFrame:
    if not keyword or not keyword.strip():
        return df
    keyword = keyword.strip()
    mask = df.astype(str).apply(
        lambda row: row.str.contains(keyword, case=False, na=False).any(),
        axis=1,
    )
    return df[mask]


def style_status_value(val):
    s = str(val).strip()
    if s in ["계약", "완료", "시공완료", "유찰", "낙찰", "진행완료", "접수완료"]:
        return "background-color: #dff0d8; color: #1b5e20; font-weight: bold;"
    if s in ["진행중", "상담중", "검토중", "운영사 변경중", "필", "상", "접수중", "시공전"]:
        return "background-color: #fff4cc; color: #7a5c00; font-weight: bold;"
    if s in ["부결", "실패", "보류", "미진행", "하", "불가", "미접수"]:
        return "background-color: #f8d7da; color: #8a1f2d; font-weight: bold;"
    return ""


def convert_number_display(value, col_name=""):
    try:
        if value is None or pd.isna(value):
            return ""

        s = str(value).strip()
        if s == "":
            return ""

        # 날짜형 값은 그대로 유지
        if re.match(r"^\d{2,4}[.\-/]\d{1,2}([.\-/]\d{1,2})?$", s):
            return s

        numeric_cols = ["수량", "판매가격", "영업수수료", "설치대수", "주차면"]
        if col_name in numeric_cols:
            s2 = s.replace("₩", "").replace("￦", "").replace(",", "").replace(" ", "")
            num = pd.to_numeric(s2, errors="coerce")

            if pd.notna(num):
                # 정수면 문자열 정수로 반환
                if float(num).is_integer():
                    return f"{int(num):,}"

                # 소수점이 실제 있을 때만 표시
                return f"{float(num):,.2f}".rstrip("0").rstrip(".")

        return s
    except Exception:
        return str(value) if value is not None else ""
        return value


def prepare_display_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    df = df.copy()
    df.columns = make_unique_columns(df.columns)

    # 통화 전용 컬럼 제거
    drop_cols = [c for c in df.columns if str(c).strip() in ["₩", "￦"]]
    if drop_cols:
        df = df.drop(columns=drop_cols, errors="ignore")

    # 완전히 비어 있는 빈컬럼 제거
    keep_cols = []
    for col in df.columns:
        col_name = str(col).strip()
        col_data = df[col]

        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]

        if col_name.startswith("빈컬럼"):
            if not col_data.astype(str).replace("", pd.NA).isna().all():
                keep_cols.append(col)
        else:
            keep_cols.append(col)

    df = df[keep_cols].copy()

    for col in df.columns:
        col_data = df[col]
        if isinstance(col_data, pd.DataFrame):
            col_data = col_data.iloc[:, 0]
        df[col] = col_data.apply(lambda x: convert_number_display(x, col))

    return df


def styled_dataframe(df: pd.DataFrame):
    df = prepare_display_df(df)

    target_cols = [
        c for c in df.columns
        if c in ["진행여부", "결과", "시공여부", "확인", "낙찰여부", "단지 반응도", "가능성", "구분", "상태", "설치유무"]
    ]

    if not target_cols:
        st.dataframe(df, use_container_width=True, height=500, hide_index=True)
        return

    styled = df.style
    for col in target_cols:
        styled = styled.map(style_status_value, subset=[col])

    st.dataframe(styled, use_container_width=True, height=500, hide_index=True)


def is_done_status(value):
    v = str(value).strip().lower()
    done_values = ["완료", "발행완료", "발행", "처리완료", "ok", "o", "y", "yes", "완", "끝"]
    return v in [x.lower() for x in done_values]


def get_d_day_label(target_date):
    today = pd.to_datetime(datetime.today().date())
    if pd.isna(target_date):
        return ""
    diff = (target_date - today).days
    if diff < 0:
        return f"D+{abs(diff)}"
    elif diff == 0:
        return "D-Day"
    else:
        return f"D-{diff}"


def make_alert_status(target_date, done_value):
    if is_done_status(done_value):
        return "완료"

    today = pd.to_datetime(datetime.today().date())
    if pd.isna(target_date):
        return "날짜없음"

    diff = (target_date - today).days
    if diff < 0:
        return "지남"
    elif diff == 0:
        return "오늘"
    elif diff <= 3:
        return "긴급"
    elif diff <= 7:
        return "임박"
    else:
        return "예정"


def style_alert_value(val):
    s = str(val).strip()
    if s == "완료":
        return "background-color: #dff0d8; color: #1b5e20; font-weight: bold;"
    if s in ["오늘", "긴급", "지남"]:
        return "background-color: #f8d7da; color: #8a1f2d; font-weight: bold;"
    if s in ["임박", "예정"]:
        return "background-color: #fff4cc; color: #7a5c00; font-weight: bold;"
    return ""


def show_alert_table(df: pd.DataFrame):
    if df.empty:
        st.info("해당 알림이 없습니다.")
        return

    styled = df.style
    if "상태표시" in df.columns:
        styled = styled.map(style_alert_value, subset=["상태표시"])
    st.dataframe(styled, use_container_width=True, hide_index=True)


# =========================================================
# 엑셀 다운로드
# =========================================================
def to_excel_bytes(data_dict):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        written_count = 0

        for sheet_name, df in data_dict.items():
            try:
                if df is None:
                    continue
                if not isinstance(df, pd.DataFrame):
                    df = pd.DataFrame(df)

                df2 = df.copy()
                if len(df2.columns) == 0 and df2.empty:
                    df2 = pd.DataFrame({"안내": ["데이터 없음"]})
                else:
                    df2.columns = [clean_text(str(col)) for col in df2.columns]
                    df2 = df2.apply(lambda col: col.map(clean_text))

                safe_sheet_name = str(sheet_name).strip() if sheet_name else f"Sheet{written_count + 1}"
                safe_sheet_name = clean_text(safe_sheet_name)[:31] or f"Sheet{written_count + 1}"
                df2.to_excel(writer, index=False, sheet_name=safe_sheet_name)
                written_count += 1
            except Exception as e:
                error_df = pd.DataFrame({"오류": [f"{sheet_name} 시트 저장 실패"], "상세": [str(e)]})
                error_df.to_excel(writer, index=False, sheet_name=f"오류{written_count + 1}"[:31])
                written_count += 1

        if written_count == 0:
            pd.DataFrame({"안내": ["저장할 데이터가 없습니다."]}).to_excel(writer, index=False, sheet_name="안내")

    output.seek(0)
    return output.getvalue()


def download_section(title, df, file_name):
    try:
        if df is None:
            st.warning(f"{title} 데이터 없음")
            return
        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame(df)
        if df.empty:
            st.warning(f"{title} 데이터 없음")
            return

        excel_data = to_excel_bytes({title: df})
        st.download_button(
            label=f"📥 {title} 다운로드",
            data=excel_data,
            file_name=f"{file_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_{file_name}_{title}",
        )
    except Exception as e:
        st.error(f"{title} 다운로드 오류: {e}")


# =========================================================
# 로고/헤더
# =========================================================
def render_header():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(base_dir, "logo.png")

    col1, col2 = st.columns([1, 5])

    with col1:
        if os.path.exists(logo_path):
            st.image(logo_path, width=100)

    with col2:
        st.markdown("## ")  # 필요하면 제목

    st.divider()

def save_vacation_log(action, target_name="", use_date="", used_days="", reason="", note=""):
    """
    연차 사용/취소/수정/재계산 로그 저장
    """
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        actor = st.session_state.get("username", "")
        if not actor:
            actor = st.session_state.get("name", "")
        if not actor:
            actor = "알수없음"

        sh = get_google_sheet()
        try:
            ws = sh.worksheet(VACATION_LOG_SHEET_NAME)
        except:
            ws = sh.add_worksheet(title=VACATION_LOG_SHEET_NAME, rows=1000, cols=8)
            ws.append_row(["기록일시", "작업자", "작업구분", "대상직원", "사용일자", "사용일수", "사유", "비고"])

        ws.append_row([
            now,
            actor,
            action,
            str(target_name),
            str(use_date),
            str(used_days),
            str(reason),
            str(note),
        ])

    except Exception as e:
        st.warning(f"연차 로그 저장 중 오류가 발생했습니다: {e}")

def save_vacation_data_to_excel(df: pd.DataFrame):
    df = df.copy()

    # ✅ 숫자 컬럼 강제 float 처리 (반차 0.5 대응)
    for col in ["발생 연차", "사용 연차", "잔여 연차"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(float)

    wb = load_workbook(VACATION_FILE_PATH)
    ws = wb[VACATION_SHEET_NAME]

    header_row = 2     # 실제 헤더 행
    start_row = 3      # 실제 데이터 시작 행

    # 엑셀 헤더 읽기
    excel_headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        excel_headers.append(str(v).strip() if v is not None else "")

    # 헤더명 -> 엑셀 컬럼번호
    col_map = {name: idx for idx, name in enumerate(excel_headers, start=1) if name}

    # df에 있는 컬럼만 엑셀에서 지우기
    last_row = max(ws.max_row, start_row + len(df) + 50)
    for r in range(start_row, last_row + 1):
        for col_name in df.columns:
            if col_name in col_map:
                ws.cell(row=r, column=col_map[col_name]).value = None

    # 헤더명 기준으로 정확히 저장
    for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for col_name in df.columns:
            if col_name in col_map:
                value = row[col_name]

                if pd.isna(value):
                    value = None

                ws.cell(row=row_idx, column=col_map[col_name]).value = value

    wb.save(VACATION_FILE_PATH)


# =========================================================
# 연차 관리 설정
# =========================================================
VACATION_BACKUP_DIR = "backup"
VACATION_FILE_PATH = "data/vacation.csv"
USE_COLS = [f"사용일{i}" for i in range(1, 70)]
VACATION_LOG_SHEET_NAME = "연차사용로그"

def to_number(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(num) else float(num)

def format_leave_number(value):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return ""
    if float(num).is_integer():
        return str(int(num))
    return str(num)

def format_display_date(value):
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except:
        return ""
    
def format_leave_date(use_date, leave_type):
    date_str = pd.to_datetime(use_date).strftime("%Y-%m-%d")
    if leave_type == "반차":
        return f"{date_str} (반차)"
    return date_str   
    
def get_target_year():
    return datetime.today().year    

def calculate_service_years(hire_date, base_date):
    years = base_date.year - hire_date.year
    if (base_date.month, base_date.day) < (hire_date.month, hire_date.day):
        years -= 1
    return max(0, years)


def calculate_anniversary_period(hire_date, target_year):
    try:
        start = date(target_year, hire_date.month, hire_date.day)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            start = date(target_year, 2, 28)
        else:
            raise

    try:
        end = date(target_year + 1, hire_date.month, hire_date.day) - timedelta(days=1)
    except ValueError:
        if hire_date.month == 2 and hire_date.day == 29:
            end = date(target_year + 1, 2, 28) - timedelta(days=1)
        else:
            raise

    return start, end


def calculate_auto_leave_days(hire_date, target_year=None):
    today = date.today()

    hire_date = pd.to_datetime(hire_date, errors="coerce")
    if pd.isna(hire_date):
        return None, None, 0, 0

    hire_date = hire_date.date()

    def safe_replace_year(d, year):
        try:
            return d.replace(year=year)
        except ValueError:
            if d.month == 2 and d.day == 29:
                return date(year, 2, 28)
            raise

    one_year_anniversary = safe_replace_year(hire_date, hire_date.year + 1)

    if today < one_year_anniversary:
        start_date = hire_date
        end_date = one_year_anniversary - timedelta(days=1)
        service_years = 0

        months_worked = (today.year - hire_date.year) * 12 + (today.month - hire_date.month)
        if today.day < hire_date.day:
            months_worked -= 1

        months_worked = max(0, min(11, months_worked))
        leave_days = float(months_worked)

    else:
        this_year_anniversary = safe_replace_year(hire_date, today.year)

        if today >= this_year_anniversary:
            start_date = this_year_anniversary
        else:
            start_date = safe_replace_year(hire_date, today.year - 1)

        end_date = safe_replace_year(start_date, start_date.year + 1) - timedelta(days=1)

        service_years = today.year - hire_date.year
        if (today.month, today.day) < (hire_date.month, hire_date.day):
            service_years -= 1

        extra_days = max(0, (service_years - 1) // 2)
        leave_days = float(min(25, 15 + extra_days))

    return start_date, end_date, service_years, leave_days


def parse_use_entry(value):
    if pd.isna(value):
        return None, None

    text = str(value).strip()
    if text == "" or text.lower() == "none":
        return None, None

    amount = 0.5 if "반차" in text else 1.0

    clean = text.replace("(반차)", "").strip()
    clean = clean.replace(".", "-").replace(" ", "")

    if len(clean) == 8 and clean[2] == "-":
        clean = "20" + clean

    parsed_date = pd.to_datetime(clean, errors="coerce")

    if pd.isna(parsed_date):
        return None, None

    return parsed_date, amount


def parse_cancel_amount(value):
    text = str(value)
    return 0.5 if "반차" in text else 1.0


def recalculate_vacation_summary(df: pd.DataFrame):
    for idx in df.index:
        total_leave = to_number(df.loc[idx, "발생 연차"])
        used_leave = 0.0

        start_date = pd.to_datetime(df.loc[idx, "기산시작일"], errors="coerce")
        end_date = pd.to_datetime(df.loc[idx, "기산종료일"], errors="coerce")

        for col in USE_COLS:
            if col not in df.columns:
                continue

            value = df.loc[idx, col]

            if pd.isna(value):
                continue

            text = str(value).strip()
            if text == "" or text.lower() == "none":
                continue

            parsed_date, amount = parse_use_entry(value)
            if parsed_date is None:
                continue

            if pd.notna(start_date) and pd.notna(end_date):
                if start_date.date() <= parsed_date.date() <= end_date.date():
                    used_leave += amount
            else:
                used_leave += amount

        remain_leave = total_leave - used_leave

        df.loc[idx, "사용 연차"] = format_leave_number(used_leave)
        df.loc[idx, "잔여 연차"] = format_leave_number(remain_leave)

    return df


def recalculate_all_vacation_data(df: pd.DataFrame):
    df = df.copy()

    for idx in df.index:
        hire_date = pd.to_datetime(df.loc[idx, "입사일"], errors="coerce")

        if pd.isna(hire_date):
            continue

        hire_date = hire_date.date()

        start_date, end_date, service_years, leave_days = calculate_auto_leave_days(hire_date)

        df.loc[idx, "기산시작일"] = str(start_date)
        df.loc[idx, "기산종료일"] = str(end_date)
        df.loc[idx, "근속년수"] = int(service_years)
        df.loc[idx, "발생 연차"] = float(leave_days)

    df = recalculate_vacation_summary(df)
    return df

def refresh_expired_vacation_rows(df: pd.DataFrame):
    """
    기산종료일이 지난 직원만 자동 갱신
    전체 재정리가 아니라 만료된 직원 행만 처리
    """
    df = df.copy()
    today = date.today()
    changed = False
    changed_names = []

    for idx in df.index:
        end_date = pd.to_datetime(df.loc[idx, "기산종료일"], errors="coerce")

        if pd.isna(end_date):
            continue

        if today > end_date.date():
            hire_date = pd.to_datetime(df.loc[idx, "입사일"], errors="coerce")

            if pd.isna(hire_date):
                continue

            start_date, new_end_date, service_years, leave_days = calculate_auto_leave_days(hire_date.date())

            df.loc[idx, "기산시작일"] = str(start_date)
            df.loc[idx, "기산종료일"] = str(new_end_date)
            df.loc[idx, "근속년수"] = int(service_years)
            df.loc[idx, "발생 연차"] = float(leave_days)

            changed = True
            changed_names.append(str(df.loc[idx, "이름"]))

    if changed:
        df = recalculate_vacation_summary(df)

    return df, changed, changed_names

def build_monthly_stats(df, target_year, target_month):
    rows = []
    total_count = 0
    total_amount = 0.0

    for _, row in df.iterrows():
        emp_name = str(row.get("이름", "")).strip()
        emp_count = 0
        emp_amount = 0.0

        for col in USE_COLS:
            value = row.get(col, None)
            parsed_date, amount = parse_use_entry(value)
            if parsed_date is None:
                continue

            if parsed_date.year == int(target_year) and parsed_date.month == int(target_month):
                emp_count += 1
                emp_amount += amount

        if emp_count > 0:
            rows.append({
                "이름": emp_name,
                "사용 건수": emp_count,
                "사용 일수": format_leave_number(emp_amount)
            })
            total_count += emp_count
            total_amount += emp_amount

    return pd.DataFrame(rows), total_count, total_amount
    

def render_employee_vacation_cards(df: pd.DataFrame):
    st.markdown(
        '<div class="erp-section-title">직원별 연차 요약 카드</div>',
        unsafe_allow_html=True
    )
    st.markdown(
        '<div class="erp-section-desc">전체 직원 기준</div>',
        unsafe_allow_html=True
    )

    if df.empty:
        st.info("표시할 직원 데이터가 없습니다.")
        return

    card_df = df.copy()

    for col in ["발생 연차", "사용 연차", "잔여 연차"]:
        if col not in card_df.columns:
            card_df[col] = 0
        card_df[col] = pd.to_numeric(card_df[col], errors="coerce").fillna(0)

    card_df["사용률"] = card_df.apply(
        lambda row: 0 if float(row["발생 연차"]) <= 0 else round((float(row["사용 연차"]) / float(row["발생 연차"])) * 100, 1),
        axis=1
    )

    card_df = card_df.sort_values(by=["잔여 연차", "사용률"], ascending=[True, False]).reset_index(drop=True)

    cols_per_row = 4

    for start in range(0, len(card_df), cols_per_row):
        row_cols = st.columns(cols_per_row)

        for i in range(cols_per_row):
            idx = start + i
            if idx >= len(card_df):
                row_cols[i].empty()
                continue

            row = card_df.iloc[idx]

            name = str(row["이름"]).strip()
            total = float(row["발생 연차"])
            used = float(row["사용 연차"])
            remain = float(row["잔여 연차"])
            rate = float(row["사용률"])

            if remain <= 0:
                status_text = "🔴 위험"
                border_color = "#ef4444"
                bg_color = "#fef2f2"
            elif remain <= 5:
                status_text = "🟡 주의"
                border_color = "#f59e0b"
                bg_color = "#fffbeb"
            else:
                status_text = "🟢 정상"
                border_color = "#22c55e"
                bg_color = "#f0fdf4"

            row_cols[i].markdown(
                f"""
                <div style="
                    border: 2px solid {border_color};
                    background: {bg_color};
                    border-radius: 14px;
                    padding: 16px;
                    margin-bottom: 12px;
                    min-height: 170px;
                ">
                    <div style="font-size:20px; font-weight:800; margin-bottom:10px;">
                        {name}
                    </div>
                    <div style="font-size:15px; line-height:1.9;">
                        • 발생 연차: <b>{format_leave_number(total)}일</b><br>
                        • 사용 연차: <b>{format_leave_number(used)}일</b><br>
                        • 잔여 연차: <b>{format_leave_number(remain)}일</b><br>
                        • 사용률: <b>{rate}%</b><br>
                        • 상태: <b>{status_text}</b>
                    </div>
                </div>
                """,
                unsafe_allow_html=True
            )

def style_remaining_leave(val):
    num = pd.to_numeric(val, errors="coerce")
    if pd.isna(num):
        return ""
    if num <= 0:
        return "background-color: #f8d7da; color: #842029; font-weight: bold;"
    elif num <= 5:
        return "background-color: #fff3cd; color: #664d03; font-weight: bold;"
    return ""

def find_first_empty_use_col(row, df_columns):
    for col in USE_COLS:
        matching_indexes = [i for i, c in enumerate(df_columns) if str(c).strip() == col]

        for col_idx in matching_indexes:
            value = row.iloc[col_idx]

            if pd.isna(value) or clean_text(value) == "" or clean_text(value).lower() == "none":
                return col_idx

    return None    

def create_backup():
    os.makedirs(VACATION_BACKUP_DIR, exist_ok=True)

    df = load_df("연차관리")
    backup_path = os.path.join(
        VACATION_BACKUP_DIR,
        f"연차관리_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    )

    df.to_csv(backup_path, index=False, encoding="utf-8-sig")
    return backup_path

def save_vacation_data(df):
    sheet_urls = get_current_sheet_urls()
    url = sheet_urls.get("연차관리", "")

    if not url:
        raise Exception("연차관리 구글시트 URL이 없습니다.")

    client = get_gsheet_client()
    sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", url).group(1)
    spreadsheet = client.open_by_key(sheet_id)
    worksheet = spreadsheet.get_worksheet(0)

    save_df = df.copy()

    if "row_id" in save_df.columns:
        save_df = save_df.drop(columns=["row_id"])

    values = worksheet.get_all_values()
    if not values:
        raise Exception("연차관리 시트에 헤더가 없습니다.")

    headers = [str(x).strip() for x in values[0]]
    headers = [h for h in headers if h != ""]

    for col in headers:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[headers].copy()
    save_df = save_df.where(pd.notnull(save_df), "")

    for col in save_df.columns:
        save_df[col] = save_df[col].apply(
            lambda x: x.strftime("%Y-%m-%d") if isinstance(x, (datetime, date, pd.Timestamp)) else str(x).strip()
        )

    rows = save_df.values.tolist()

    worksheet.update("A2", rows, value_input_option="USER_ENTERED")

    old_rows = max(0, len(values) - 1)
    new_rows = len(rows)

    if old_rows > new_rows:
        blank_rows = old_rows - new_rows
        clear_start = new_rows + 2
        empty_rows = [[""] * len(headers) for _ in range(blank_rows)]
        worksheet.update(f"A{clear_start}", empty_rows)

    load_google_sheet_data.clear() 

# =========================================================
# 시공 일정 시스템
# =========================================================
SCHEDULE_SHEET_NAME = "시공일정"
EXPECTED_COLUMNS = ["날짜", "상품구분", "설치현장", "시공담당", "수량", "비고", "상태", "완료일"]

def get_schedule_sheet():
    url = get_current_sheet_urls().get("시공일정")

    if not url:
        st.error("시공일정 URL이 없습니다.")
        return None

    sheet_id = re.search(r"/d/([a-zA-Z0-9-_]+)", url).group(1)

    gid_match = re.search(r"gid=(\d+)", url)
    gid = int(gid_match.group(1)) if gid_match else 0

    client = get_gsheet_client()
    spreadsheet = client.open_by_key(sheet_id)

    sheet = spreadsheet.get_worksheet_by_id(gid)

    if sheet is None:
        st.error(f"시공일정 시트를 찾지 못했습니다. gid={gid}")
        return None

    return sheet

def append_schedule_data(new_row_df):
    sheet = get_schedule_sheet()

    save_df = new_row_df.copy()

    # 🚨 상품구분 컬럼 없으면 추가 저장 금지
    if "상품구분" not in save_df.columns:
        st.error("추가 저장 데이터에 상품구분 컬럼이 없습니다.")
        return

    for col in EXPECTED_COLUMNS:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[EXPECTED_COLUMNS].fillna("")

    rows = save_df.astype(str).values.tolist()

    sheet.append_rows(rows, value_input_option="USER_ENTERED")

    load_schedule_data.clear()   

def ensure_schedule_sheet_header(sheet):
    values = sheet.get_all_values()

    if not values:
        st.error("시공일정 시트가 비어 있습니다. 헤더를 먼저 확인해주세요.")
        return False

    header = [str(x).strip() for x in values[0]]

    if header[:len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        st.warning(
            "시공일정 구글시트 헤더가 기준과 다릅니다. "
            "앱에서는 저장하지 않고 읽기만 제한합니다."
        )
        return False

    return True

@st.cache_data(ttl=300)
def load_schedule_data():
    sheet = get_schedule_sheet()

    values = sheet.get_all_values()

    if not values or len(values) < 2:
        return pd.DataFrame(columns=EXPECTED_COLUMNS)

    header = [str(x).strip() for x in values[0]]

    # ✅ 헤더가 정상 아니면 앱에서 멈추지 않고 경고만 표시
    if header[:len(EXPECTED_COLUMNS)] != EXPECTED_COLUMNS:
        st.error(
            "시공일정 구글시트 헤더가 기준과 다릅니다. "
            "데이터 보호를 위해 읽기/저장을 중단합니다."
        )
        return pd.DataFrame(columns=EXPECTED_COLUMNS)

    rows = values[1:]
    df = pd.DataFrame(rows, columns=EXPECTED_COLUMNS)

    df = df.fillna("")

    # 수량 숫자 보정
    df["수량"] = (
        pd.to_numeric(df["수량"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # 타입 보정
    df["날짜"] = df["날짜"].astype(str)
    df["완료일"] = df["완료일"].astype(str)
    df["상태"] = df["상태"].astype(str).replace("", "진행중")

    return df


def save_schedule_data(df, sheet=None):
    if sheet is None:
        sheet = get_schedule_sheet()

    save_df = df.copy()
    # 🚨 저장 원본에 상품구분 컬럼 없으면 저장 금지
    if "상품구분" not in save_df.columns:
        st.error("저장 원본에 상품구분 컬럼이 없습니다. 저장을 중단합니다.")
        return
    st.warning(f"저장 직전 시공일정 행 수: {len(save_df)}")
    st.warning(f"저장 직전 컬럼: {list(save_df.columns)}")

    if "상품구분" in save_df.columns:
        st.warning(
            "저장 직전 상품구분 값: "
            + str(save_df["상품구분"].astype(str).str.strip().unique().tolist())
        )
    else:
        st.error("저장 직전 상품구분 컬럼이 없습니다.")

    for col in EXPECTED_COLUMNS:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[EXPECTED_COLUMNS].fillna("")
    if list(save_df.columns) != EXPECTED_COLUMNS:
        st.error("컬럼 구조 이상 - 저장 중단")
        return
    # ✅ 시공일정 전체 빈 데이터 저장 방지
    if len(save_df) == 0:
        st.error("시공일정 데이터가 0건입니다. 저장을 중단합니다.")
        return

    # ✅ 상품구분 전체 공란 저장 방지
    if (
        len(save_df) > 0
        and save_df["상품구분"].astype(str).str.strip().eq("").all()
    ):
        st.error("상품구분이 모두 비어 있어 저장을 중단합니다.")
        return
    
    # ✅ 필수값 정리
    save_df["날짜"] = save_df["날짜"].astype(str).str.strip()
    save_df["상품구분"] = save_df["상품구분"].astype(str).str.strip()
    save_df["설치현장"] = save_df["설치현장"].astype(str).str.strip()
    save_df["시공담당"] = save_df["시공담당"].astype(str).str.strip()
    save_df["상태"] = save_df["상태"].astype(str).str.strip()

    # ✅ 빈 행 저장 방지
    save_df = save_df[save_df["날짜"] != ""].copy()
    save_df = save_df[save_df["설치현장"] != ""].copy()
    save_df = save_df[save_df["시공담당"] != ""].copy()

    # ✅ 상태값 보정
    save_df.loc[~save_df["상태"].isin(["진행중", "완료"]), "상태"] = "진행중"
    save_df["수량"] = pd.to_numeric(save_df["수량"], errors="coerce").fillna(0).astype(int)

    # 기존 시트 데이터 길이 확인
    rows = [save_df.columns.tolist()] + save_df.astype(str).values.tolist()

    sheet.clear()

    sheet.update(
        "A1",
        rows,
        value_input_option="USER_ENTERED"
    )

    load_schedule_data.clear()

# =========================================================
# 업무일정 구글시트 load/save
# =========================================================
@st.cache_data(ttl=300)
def load_work_schedule_data():
    try:
        client = get_gsheet_client()
        spreadsheet = client.open_by_url(WORK_SCHEDULE_SHEET_URL)
        ws = spreadsheet.worksheet(WORK_SCHEDULE_SHEET_NAME)

        records = ws.get_all_records()
        df = pd.DataFrame(records)

        if df.empty:
            df = pd.DataFrame(columns=WORK_SCHEDULE_COLUMNS)

        for col in WORK_SCHEDULE_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        return df[WORK_SCHEDULE_COLUMNS].copy()

    except Exception as e:
        st.warning(f"업무일정 구글시트 불러오기 실패: {e}")
        return pd.DataFrame(columns=WORK_SCHEDULE_COLUMNS)


def save_work_schedule_data(df):
    try:
        client = get_gsheet_client()
        spreadsheet = client.open_by_url(WORK_SCHEDULE_SHEET_URL)
        ws = spreadsheet.worksheet(WORK_SCHEDULE_SHEET_NAME)

        save_df = df.copy()

        for col in WORK_SCHEDULE_COLUMNS:
            if col not in save_df.columns:
                save_df[col] = ""

        save_df = save_df[WORK_SCHEDULE_COLUMNS].fillna("")

        ws.clear()
        ws.update([WORK_SCHEDULE_COLUMNS] + save_df.values.tolist())

        return True

    except Exception as e:
        st.error(f"업무일정 구글시트 저장 실패: {e}")
        return False


SCHEDULE_LOG_COLUMNS = ["시간", "사용자", "사업", "작업", "설치현장", "시공담당", "비고"]

def find_original_schedule_index(full_df, target_row):
    full_df = full_df[EXPECTED_COLUMNS].copy().fillna("")

    for col in EXPECTED_COLUMNS:
        full_df[col] = full_df[col].astype(str).str.strip()

    mask = (
        (full_df["날짜"] == str(target_row["날짜"]).strip()) &
        (full_df["상품구분"] == str(target_row["상품구분"]).strip()) &
        (full_df["설치현장"] == str(target_row["설치현장"]).strip()) &
        (full_df["시공담당"] == str(target_row["시공담당"]).strip()) &
        (full_df["수량"].astype(str) == str(target_row["수량"]).strip()) &
        (full_df["비고"] == str(target_row["비고"]).strip()) &
        (full_df["상태"] == str(target_row["상태"]).strip())
    )

    matched = full_df[mask]

    if matched.empty:
        return None

    return matched.index[0]

def save_schedule_log(action, site="", manager="", note=""):
    try:
        log_path = os.path.join(DATA_DIR, "schedule_work_log.csv")

        user = str(st.session_state.get("display_name", "")).strip()
        business = str(st.session_state.get("business", "")).strip()

        new_log = pd.DataFrame([{
            "시간": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "사용자": user,
            "사업": business,
            "작업": action,
            "설치현장": site,
            "시공담당": manager,
            "비고": note,
        }])

        if os.path.exists(log_path):
            old_log = pd.read_csv(log_path, encoding="utf-8-sig").fillna("")
            log_df = pd.concat([old_log, new_log], ignore_index=True)
        else:
            log_df = new_log

        log_df = log_df[SCHEDULE_LOG_COLUMNS]
        log_df.to_csv(log_path, index=False, encoding="utf-8-sig")

    except Exception as e:
        st.warning(f"작업 로그 저장 실패: {e}")    

def vacation_page():

    st.markdown('<div class="erp-page-title">윤우테크 연차 관리 프로그램</div>', unsafe_allow_html=True)
    st.markdown('<div class="erp-page-desc">연차 관리 및 현황 확인 프로그램입니다.</div>', unsafe_allow_html=True)

    login_id = str(
        st.session_state.get("user_id")
        or st.session_state.get("username")
        or ""
    ).strip()

    login_role = str(
        st.session_state.get("role")
        or st.session_state.get("user_role")
        or st.session_state.get("권한")
        or ""
    ).strip()

    is_admin = login_role == "관리자"

    try:
        df = load_df("연차관리")

        # ✅ 기산종료일 지난 직원만 자동 갱신
        df, vacation_changed, changed_names = refresh_expired_vacation_rows(df)

        if vacation_changed:
            backup_file = create_backup()
            save_vacation_data(df)
            load_google_sheet_data.clear()
            st.info(
                f"기산기간이 지난 직원 연차가 자동 갱신되었습니다: "
                f"{', '.join(changed_names)} / 백업: {backup_file}"
            )
            st.rerun()

        df = apply_role_filter(df)

        users = load_users_from_gsheet()

        if not users:
            st.error("사용자관리 정보를 불러오지 못했습니다.")
            return

        if login_id not in users:
            st.error(f"사용자관리 시트에서 로그인 ID '{login_id}'를 찾지 못했습니다.")
            return

        login_name = str(users[login_id].get("name", "")).strip()

        if is_admin:
            df_view = df.copy()
        else:
            df_view = df[df["이름"] == login_name].copy()

    except Exception as e:
        st.error(f"연차 파일을 불러오지 못했습니다: {e}")
        return

    # =====================================================
    # 관리 도구
    # =====================================================
    if is_admin:
        st.markdown(
            '<div class="erp-section-title">🛠 관리 도구</div>',
            unsafe_allow_html=True
        )

        tool_col1, tool_col2, tool_col3 = st.columns(3)

        with tool_col1:
            if st.button("💾 지금 백업하기", use_container_width=True, key="vac_backup_btn_unique"):
                backup_file = create_backup()
                st.success(f"백업 완료: {backup_file}")

        with tool_col2:
            confirm_recalc = st.checkbox("정말 전체 연차를 재정리하고 구글시트에 저장합니다. 백업 후에만 체크하세요.")

            if st.button("📊 연차 수치 재정리", use_container_width=True, key="vac_recalc_btn_unique"):
                if not confirm_recalc:
                    st.warning("체크 후 실행하세요.")
                else:
                    df = recalculate_all_vacation_data(df)
                    save_vacation_data(df)
                    load_google_sheet_data.clear()
                    st.success("전체 재계산 완료")
                    st.rerun()

        with tool_col3:
            if os.path.exists(VACATION_BACKUP_DIR):
                backup_files = sorted(os.listdir(VACATION_BACKUP_DIR), reverse=True)
                st.write(f"백업 파일 수: {len(backup_files)}")
            else:
                st.write("백업 파일 수: 0")

    # =====================================================
    # 직원 선택
    # =====================================================
    st.markdown(
        '<div class="erp-section-title">👥 직원 선택</div>',
        unsafe_allow_html=True
    )

    names = sorted(df_view["이름"].dropna().astype(str).unique().tolist())

    if is_admin:
        search_name = st.text_input("직원 검색", placeholder="이름을 입력하세요", key="vac_search_name_unique")

        if search_name:
            filtered_names = [n for n in names if search_name.strip().lower() in n.lower()]
        else:
            filtered_names = names

        if not filtered_names:
            st.warning("검색 결과가 없습니다.")
            return

        selected_name = st.selectbox("직원 선택", filtered_names, key="vac_selected_name_unique")
    else:
        selected_name = login_name
        st.info(f"본인 연차만 조회됩니다: {selected_name}")

    employee_rows = df_view[df_view["이름"] == selected_name]

    if employee_rows.empty:
        st.error(f"연차관리 시트에서 '{selected_name}' 직원을 찾지 못했습니다.")
        return

    employee = employee_rows.iloc[0]

    # =====================================================
    # 현재 연차 현황
    # =====================================================
    st.markdown(
        '<div class="erp-section-title">📌 현재 연차 현황</div>',
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns(3)

    total = to_number(employee["발생 연차"])
    used = to_number(employee["사용 연차"])
    remain = to_number(employee["잔여 연차"])

    with col1:
        ui_card("총 연차", format_leave_number(total), "발생 연차")

    with col2:
        ui_card("사용 연차", format_leave_number(used), "사용 완료")

    with col3:
        ui_card("잔여 연차", format_leave_number(remain), "현재 잔여")

    if remain <= 0:
        st.error("잔여 연차가 없습니다.")
    elif remain <= 5:
        st.warning("잔여 연차가 5일 이하입니다.")
    else:
        st.success("잔여 연차가 충분합니다.")

    if is_admin:
        if st.button("🔄 선택 직원 연차 다시 계산", use_container_width=True, key="vac_recalc_selected_btn"):

            # ✅ 선택 직원의 실제 행 위치를 숫자로 찾기
            match_positions = [
                i for i, name in enumerate(df["이름"].astype(str).str.strip().tolist())
                if name == str(selected_name).strip()
            ]

            if not match_positions:
                st.error("선택한 직원을 찾지 못했습니다.")
                st.stop()

            row_pos = match_positions[0]

            # ✅ 컬럼 위치를 숫자로 찾기
            used_col_pos = list(df.columns).index("사용 연차")
            remain_col_pos = list(df.columns).index("잔여 연차")
            total_col_pos = list(df.columns).index("발생 연차")
            start_col_pos = list(df.columns).index("기산시작일")
            end_col_pos = list(df.columns).index("기산종료일")

            total_leave = to_number(df.iloc[row_pos, total_col_pos])
            used_leave = 0.0

            start_date = pd.to_datetime(df.iloc[row_pos, start_col_pos], errors="coerce")
            end_date = pd.to_datetime(df.iloc[row_pos, end_col_pos], errors="coerce")

            for col in USE_COLS:
                if col not in df.columns:
                    continue

                col_pos = list(df.columns).index(col)
                value = df.iloc[row_pos, col_pos]

                if pd.isna(value):
                    continue

                text = str(value).strip()
                if text == "" or text.lower() == "none":
                    continue

                parsed_date, amount = parse_use_entry(value)

                if parsed_date is None:
                    continue

                if pd.notna(start_date) and pd.notna(end_date):
                    if start_date.date() <= parsed_date.date() <= end_date.date():
                        used_leave += amount
                else:
                    used_leave += amount

            remain_leave = total_leave - used_leave

            # ✅ 배포앱 안전 처리: df 전체를 object 타입으로 변경
            df = df.astype(object)

            # ✅ 선택 직원 값 저장
            df.iloc[row_pos, used_col_pos] = format_leave_number(used_leave)
            df.iloc[row_pos, remain_col_pos] = format_leave_number(remain_leave)

            save_vacation_log(
                action="재계산",
                target_name=str(selected_name),
                use_date="",
                used_days="",
                reason="",
                note="선택 직원 연차 다시 계산"
            )
            load_google_sheet_data.clear()
            st.success(f"{selected_name} 연차가 다시 계산되었습니다.")
            st.rerun()

    # =====================================================
    # 관리자: 직원별 요약 카드
    # =====================================================
    if is_admin:
        with st.expander("직원별 연차 요약 카드 (전체 직원)", expanded=False):
            render_employee_vacation_cards(df)

    # =====================================================
    # 연차 사용 입력
    # =====================================================
    if is_admin:    
        with st.expander("📝 연차 사용 입력", expanded=False):
            use_date = st.date_input("사용 날짜 선택", datetime.today(), key="vac_use_date_unique")
            leave_type = st.radio("사용 종류 선택", ["연차", "반차"], horizontal=True, key="vac_leave_type_unique")
            leave_amount = 1.0 if leave_type == "연차" else 0.5

            st.write(f"선택된 사용값: **{format_leave_number(leave_amount)}일**")

            btn_col1, btn_col2 = st.columns(2)

            with btn_col1:
                register_btn = st.button("등록하기", type="primary", use_container_width=True, key="vac_register_btn_unique")

            with btn_col2:
                preview_btn = st.button("미리 확인", use_container_width=True, key="vac_preview_btn_unique")

            if preview_btn:
                expected_used = used + leave_amount
                expected_remain = total - expected_used
                st.info(
                    f"{selected_name} / {use_date.strftime('%Y-%m-%d')} / {leave_type} 등록 시 "
                    f"사용 연차 {format_leave_number(expected_used)}, 잔여 연차 {format_leave_number(expected_remain)}"
                )

            if register_btn:
                idx = df[df["이름"] == selected_name].index[0]

                current_total = float(to_number(df.loc[idx, "발생 연차"]))
                current_used = float(to_number(df.loc[idx, "사용 연차"]))
                current_remain = float(to_number(df.loc[idx, "잔여 연차"]))

                if current_remain < leave_amount:
                    st.error("잔여 연차가 부족합니다.")
                else:
                    row_pos = df.index.get_loc(idx)
                    empty_col_idx = find_first_empty_use_col(df.iloc[row_pos], df.columns)

                    if empty_col_idx is None:
                        st.error("사용일 칸이 모두 찼습니다. 사용일1~사용일30을 확인해주세요.")
                    else:

                        # ✅ 기산기간 밖 연차 입력 차단
                        start_date = pd.to_datetime(df.iloc[row_pos]["기산시작일"], errors="coerce")
                        end_date = pd.to_datetime(df.iloc[row_pos]["기산종료일"], errors="coerce")

                        if pd.notna(start_date) and pd.notna(end_date):
                            if use_date < start_date.date() or use_date > end_date.date():
                                st.error(
                                    f"⚠️ 기산기간 외 연차입니다. "
                                    f"이 직원의 기산기간은 {start_date.date()} ~ {end_date.date()} 입니다."
                                )
                                st.stop()

                        df.iat[row_pos, empty_col_idx] = format_leave_date(use_date, leave_type)

                        target_idx = df.index[row_pos]

                        target_one = df.loc[[target_idx]].copy()
                        target_one = recalculate_vacation_summary(target_one)

                        df.loc[target_idx, "사용 연차"] = target_one.loc[target_idx, "사용 연차"]
                        df.loc[target_idx, "잔여 연차"] = target_one.loc[target_idx, "잔여 연차"]

                        save_vacation_data(df)

                        save_vacation_log(
                            action="등록",
                            target_name=str(selected_name),
                            use_date=str(use_date),
                            used_days=str(leave_amount),
                            reason=str(leave_type),
                            note="연차 사용 등록"
                        )

                        load_google_sheet_data.clear()
                        st.success("연차 등록 완료!")
                        st.rerun()

    # =====================================================
    # 선택 직원 사용일 내역
    # =====================================================
    with st.expander("🗂️ 선택 직원 사용일 내역", expanded=False):
        use_list = []

        for col in USE_COLS:
            value = employee.get(col, None)
            display_value = ""

            if pd.notna(value) and clean_text(value) != "" and clean_text(value).lower() != "none":
                display_value = format_display_date(value)

                if display_value == "":
                    display_value = clean_text(value)

            use_list.append({
                "구분": col,
                "사용내역": display_value
            })

        use_df = pd.DataFrame(use_list)
        use_df = use_df[use_df["사용내역"] != ""]

        if not use_df.empty:
            st.dataframe(use_df, use_container_width=True)
        else:
            st.info("등록된 사용일이 없습니다.")

    # =====================================================
    # 연차 취소
    # =====================================================
    if is_admin:
        with st.expander("↩️ 연차 취소", expanded=False):
            use_list = []

            for col in USE_COLS:
                value = employee.get(col, None)
                display_value = ""

                if pd.notna(value) and clean_text(value) != "" and clean_text(value).lower() != "none":
                    display_value = format_display_date(value)

                    if display_value == "":
                        display_value = clean_text(value)

                use_list.append({
                    "구분": col,
                    "사용내역": display_value
                })

            use_df = pd.DataFrame(use_list)
            use_df = use_df[use_df["사용내역"] != ""]

            if not use_df.empty:
                cancel_options = [f"{row['구분']} | {row['사용내역']}" for _, row in use_df.iterrows()]
                selected_cancel = st.selectbox("취소할 사용일 선택", cancel_options, key="vac_cancel_select_unique")

                if st.button("선택 사용일 취소", use_container_width=True, key="vac_cancel_btn_unique"):
                    idx = df[df["이름"] == selected_name].index[0]

                    selected_col = selected_cancel.split("|")[0].strip()
                    selected_value = df.loc[idx, selected_col]

                    cancel_amount = parse_cancel_amount(selected_value)

                    current_total = to_number(df.loc[idx, "발생 연차"])
                    current_used = to_number(df.loc[idx, "사용 연차"])

                    new_used = max(0, current_used - cancel_amount)
                    new_remain = current_total - new_used

                    df = df.astype(object)

                    df.loc[idx, selected_col] = ""
                    df.loc[idx, "사용 연차"] = format_leave_number(new_used)
                    df.loc[idx, "잔여 연차"] = format_leave_number(new_remain)

                    save_vacation_data(df)

                    save_vacation_log(
                        action="취소",
                        target_name=str(selected_name),
                        use_date="",
                        used_days="",
                        reason="연차 사용 취소",
                        note="선택 사용일 취소"
                    )

                    load_google_sheet_data.clear()
                    st.success("연차 취소 완료!")
                    st.rerun()
            else:
                st.info("취소할 사용일이 없습니다.")

    # =====================================================
    # 관리자 전용: 직원 관리
    # =====================================================
    if is_admin:
        with st.expander("📁 직원 관리", expanded=False):

            st.markdown("## ➕ 직원 추가")

            with st.form("add_employee_form_unique"):
                new_name = st.text_input("직원 이름", key="new_employee_name_unique")
                new_hire_date = st.date_input("입사일", value=date.today(), key="new_employee_hire_date_unique")

                preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
                    new_hire_date,
                    get_target_year()
                )

                st.info(
                    f"자동 계산 결과\n\n"
                    f"- 기산시작일: {preview_start}\n"
                    f"- 기산종료일: {preview_end}\n"
                    f"- 근속년수: {preview_service_years}\n"
                    f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
                )

                submit_add_employee = st.form_submit_button("직원 추가하기")

                if submit_add_employee:
                    new_name = new_name.strip()

                    if new_name == "":
                        st.error("직원 이름을 입력해주세요.")
                    elif new_name in df["이름"].astype(str).tolist():
                        st.error("이미 등록된 직원입니다.")
                    else:
                        hire_date = pd.to_datetime(new_hire_date).date()
                        start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                            hire_date,
                            get_target_year()
                        )

                        new_row = {
                            "이름": new_name,
                            "입사일": pd.to_datetime(hire_date),
                            "기산시작일": pd.to_datetime(start_date),
                            "기산종료일": pd.to_datetime(end_date),
                            "근속년수": service_years,
                            "발생 연차": float(auto_leave_days),
                            "사용 연차": 0.0,
                            "잔여 연차": float(auto_leave_days),
                        }

                        for col in USE_COLS:
                            new_row[col] = None

                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

                        save_vacation_data(df)
                        load_google_sheet_data.clear()
                        st.success("직원 추가 완료!")
                        st.rerun()

            st.markdown("---")
            st.markdown("## ✏️ 직원 수정")

            edit_name = st.selectbox("수정할 직원 선택", names, key="edit_employee_select_unique")
            edit_employee = df[df["이름"] == edit_name].iloc[0]

            default_hire_date = pd.to_datetime(edit_employee["입사일"], errors="coerce")
            if pd.isna(default_hire_date):
                default_hire_date = pd.Timestamp(date.today())

            with st.form("edit_employee_form_unique"):
                edited_name = st.text_input("직원 이름 수정", value=str(edit_employee["이름"]), key="edited_name_unique")
                edited_hire_date = st.date_input("입사일 수정", value=default_hire_date.date(), key="edited_hire_date_unique")
                edited_used_leave = st.number_input(
                    "사용 연차 수정",
                    min_value=0.0,
                    step=0.5,
                    value=float(to_number(edit_employee["사용 연차"])),
                    key="edited_used_leave_unique"
                )

                preview_start, preview_end, preview_service_years, preview_leave_days = calculate_auto_leave_days(
                    pd.to_datetime(edited_hire_date).date(),
                    get_target_year()
                )

                st.info(
                    f"자동 계산 결과\n\n"
                    f"- 기산시작일: {preview_start}\n"
                    f"- 기산종료일: {preview_end}\n"
                    f"- 근속년수: {preview_service_years}\n"
                    f"- 발생 연차: {format_leave_number(preview_leave_days)}일"
                )

                submit_edit_employee = st.form_submit_button("직원 정보 수정하기")

                if submit_edit_employee:
                    edited_name = edited_name.strip()

                    if edited_name == "":
                        st.error("직원 이름을 입력해주세요.")
                    else:
                        duplicate_names = [n for n in df["이름"].astype(str).tolist() if n != edit_name]

                        if edited_name in duplicate_names:
                            st.error("같은 이름의 직원이 이미 있습니다.")
                        else:
                            idx = df[df["이름"] == edit_name].index[0]

                            hire_date = pd.to_datetime(edited_hire_date).date()
                            start_date, end_date, service_years, auto_leave_days = calculate_auto_leave_days(
                                hire_date,
                                get_target_year()
                            )

                            new_total = float(auto_leave_days)
                            new_used = float(edited_used_leave)
                            new_remain = new_total - new_used

                            if new_remain < 0:
                                st.error("사용 연차가 발생 연차보다 클 수 없습니다.")
                            else:
                                if "근속년수" not in df.columns:
                                    df["근속년수"] = ""

                                df["근속년수"] = pd.to_numeric(df["근속년수"], errors="coerce").fillna(0)
                                df["근속년수"] = df["근속년수"].astype(int)

                                row_pos = df.index.get_loc(idx)

                                df.loc[idx, "이름"] = edited_name
                                df.loc[idx, "입사일"] = str(hire_date)
                                df.loc[idx, "기산시작일"] = str(start_date)
                                df.loc[idx, "기산종료일"] = str(end_date)
                                df.loc[idx, "발생 연차"] = float(new_total)
                                df.loc[idx, "사용 연차"] = float(new_used)
                                df.loc[idx, "잔여 연차"] = float(new_remain)

                                service_col_pos = list(df.columns).index("근속년수")
                                df.iat[row_pos, service_col_pos] = int(service_years)

                                save_vacation_data(df)
                                load_google_sheet_data.clear()
                                st.success("직원 정보 수정 완료!")
                                st.rerun()

            st.markdown("---")
            st.markdown("## 🗑️ 직원 삭제")

            delete_name = st.selectbox("삭제할 직원 선택", names, key="delete_employee_select_unique")
            confirm_delete = st.checkbox("정말 삭제합니다. 되돌리기 어렵습니다.", key="vac_confirm_delete_unique")

            if st.button("선택 직원 삭제", use_container_width=True, key="vac_delete_btn_unique"):
                if not confirm_delete:
                    st.warning("삭제 확인 체크를 먼저 해주세요.")
                else:
                    before_count = len(df)
                    df = df[df["이름"].astype(str) != str(delete_name)].copy()
                    after_count = len(df)

                    if before_count == after_count:
                        st.error("삭제할 직원을 찾지 못했습니다.")
                    else:
                        save_vacation_data(df)
                        load_google_sheet_data.clear()
                        st.success(f"{delete_name} 직원 삭제 완료!")
                        st.rerun()

    # =====================================================
    # 월별 연차 통계
    # 관리자: 전체 / 직원: 본인만
    # =====================================================
    with st.expander("📅 월별 연차 통계", expanded=False):

        if is_admin:
            stat_base_df = df.copy()
        else:
            stat_base_df = df[df["이름"] == login_name].copy()
        stat_col1, stat_col2 = st.columns(2)
        with stat_col1:
            stat_year = st.number_input(
                "조회 연도",
                min_value=2020,
                max_value=2100,
                value=get_target_year(),
                step=1,
                key="vac_stat_year_unique"
            )

        with stat_col2:
            stat_month = st.selectbox(
                "조회 월",
                list(range(1, 13)),
                index=max(0, datetime.today().month - 1),
                key="vac_stat_month_unique"
            )

        monthly_df, monthly_count, monthly_amount = build_monthly_stats(
            stat_base_df, int(stat_year), int(stat_month)
        )

        metric_col1, metric_col2 = st.columns(2)

        with metric_col1:
            ui_card("해당 월 사용 건수", monthly_count, "사용 횟수")

        with metric_col2:
            ui_card("해당 월 총 사용일수", format_leave_number(monthly_amount), "사용 연차")

        if not monthly_df.empty:
            st.dataframe(monthly_df, use_container_width=True)
        else:
            st.info("해당 월 사용 내역이 없습니다.")

    # =====================================================
    # 관리자 전용: 전체 연차 현황
    # =====================================================
    with st.expander("📋 전체 연차 현황", expanded=False):

        if is_admin:
            display_df = df.copy()
        else:
            display_df = df[df["이름"] == login_name].copy()

        for col in ["입사일", "기산시작일", "기산종료일"]:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(format_display_date)

        for col in USE_COLS:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(
                    lambda x: (
                        format_display_date(x)
                        if format_display_date(x) != ""
                        else clean_text(x)
                    )
                )

        for col in ["발생 연차", "사용 연차", "잔여 연차"]:
            if col in display_df.columns:
                display_df[col] = display_df[col].apply(format_leave_number)

        basic_cols = [
            "이름", "입사일", "기산시작일", "기산종료일",
            "근속년수", "발생 연차", "사용 연차", "잔여 연차"
        ]

        use_cols = USE_COLS

        basic_cols = [col for col in basic_cols if col in display_df.columns]
        use_cols = [col for col in use_cols if col in display_df.columns]

        st.markdown(
            '<div class="erp-section-title">📊 기본 연차 정보</div>',
            unsafe_allow_html=True
        )

        basic_df = display_df[basic_cols].copy()

        if "잔여 연차" in basic_df.columns:
            styled_basic_df = basic_df.style.map(
                style_remaining_leave,
                subset=["잔여 연차"]
            )
            st.dataframe(styled_basic_df, use_container_width=True, height=400)
        else:
            st.dataframe(basic_df, use_container_width=True, height=400)

        st.markdown(
            '<div class="erp-section-title">🧾 연차 사용 이력</div>',
            unsafe_allow_html=True
        )

        if use_cols:
            use_df = display_df[["이름"] + use_cols].set_index("이름")
            st.dataframe(use_df, use_container_width=False, height=600)
        else:
            st.info("사용일 컬럼이 없습니다.")

    # =====================================================
    # 엑셀 다운로드
    # =====================================================
    with st.expander("⬇️ 엑셀 다운로드", expanded=False):
        download_df = load_df("연차관리")
        excel_data = to_excel_bytes({"연차관리": download_df})

        st.download_button(
            label="엑셀 다운로드",
            data=excel_data,
            file_name=f"연차관리_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="vac_download_btn_unique"
        )           

# =========================================================
# 5. 실사 관리 시스템
# =========================================================
INSPECTION_SHEET_NAME = "실사관리"
INSPECTION_COLUMNS = [
    "요청일",
    "운영사",
    "현장명",
    "현장주소",
    "현장연락처",
    "주차면수",
    "상품구분",
    "환경부",
    "자투",
    "신규설치수량",
    "기설치수량",
    "영업담당자",
    "영업담당연락처",
    "요청내용",
    "비고",
    "첨부파일명",
    "첨부파일링크",
    "실사담당자",
    "실사예정일",
    "실사완료일",
    "진행상태",
    "실사결과",
    "특이사항",
    "후속조치",
    "계약여부",
    "계약일",
    "계약수량",
    "계약금액",
    "미계약사유"
]

INSPECTION_STATUS_OPTIONS = [
    "요청접수",
    "담당자배정",
    "일정확정",
    "실사진행",
    "실사완료",
    "계약완료",
    "미계약종결"
]

PRODUCT_OPTIONS = ["아이센서", "전기차충전기", "이전설치"]
ENV_OPTIONS = ["", "대상", "비대상"]
JATU_OPTIONS = ["", "있음", "없음"]
CONTRACT_OPTIONS = ["대기", "계약", "미계약"]


def get_inspection_sheet():
    client = get_gsheet_client()
    spreadsheet = client.open(INSPECTION_SHEET_NAME)
    worksheet = spreadsheet.worksheet("실사복구")
    return worksheet


def safe_int(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return default
    return int(num)


def show_inspection_flash():
    msg = st.session_state.pop("inspection_flash", "")
    msg_type = st.session_state.pop("inspection_flash_type", "success")


    if msg:
        if msg_type == "success":
            st.success(msg)
        elif msg_type == "warning":
            st.warning(msg)
        elif msg_type == "error":
            st.error(msg)
        else:
            st.info(msg)


def set_inspection_flash(msg, msg_type="success"):
    st.session_state["inspection_flash"] = msg
    st.session_state["inspection_flash_type"] = msg_type


def normalize_inspection_df(df):
    # 컬럼 없으면 자동 생성
    for col in INSPECTION_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # 컬럼 순서 맞추기
    df = df.reindex(columns=INSPECTION_COLUMNS, fill_value="").copy()

    # 숫자 컬럼 (없는 경우도 대비)
    int_cols = ["주차면수", "신규설치수량", "기설치수량", "계약수량"]
    for col in int_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    # 금액
    if "계약금액" in df.columns:
        df["계약금액"] = pd.to_numeric(df["계약금액"], errors="coerce").fillna(0)

    # 날짜
    date_cols = ["요청일", "실사예정일", "실사완료일", "계약일"]
    for col in date_cols:
        if col in df.columns:
            df[col] = df[col].astype(str)

    # 상태값
    if "진행상태" in df.columns:
        df["진행상태"] = df["진행상태"].astype(str).replace("", "요청접수")

    if "계약여부" in df.columns:
        df["계약여부"] = df["계약여부"].astype(str).replace("", "대기")

    return df

def detect_inspection_duplicates(df):
    keys = ["요청일", "상품구분", "현장명", "현장주소", "영업담당자"]
    existing_keys = [c for c in keys if c in df.columns]

    if not existing_keys:
        return pd.DataFrame()

    dup_df = df[df.duplicated(subset=existing_keys, keep=False)].copy()
    return dup_df

def find_original_inspection_index(full_df, target_row):
    full_df = full_df[INSPECTION_COLUMNS].copy().fillna("")

    for col in INSPECTION_COLUMNS:
        full_df[col] = full_df[col].astype(str).str.strip()

    mask = (
        (full_df["요청일"] == str(target_row["요청일"]).strip()) &
        (full_df["상품구분"] == str(target_row["상품구분"]).strip()) &
        (full_df["현장명"] == str(target_row["현장명"]).strip()) &
        (full_df["현장주소"] == str(target_row["현장주소"]).strip()) &
        (full_df["영업담당자"] == str(target_row["영업담당자"]).strip())
    )

    matched = full_df[mask]

    if matched.empty:
        return None

    return matched.index[0]

# =========================================================
# 차량관리
# =========================================================

VEHICLE_SHEET_NAME = "차량관리"
VEHICLE_REPAIR_SHEET_NAME = "차량정비이력"

VEHICLE_COLUMNS = [
    "차량명", "소유자", "소유형태", "유종", "차종", "모델명", "연식", "차량번호",
    "보험회사", "보험종류", "보험기간", "보험금액", "차량상태", "비고"
]

REPAIR_COLUMNS = [
    "차량번호", "수리일자", "수리내역", "금액", "비고"
]
VEHICLE_SPREADSHEET_ID = "1OBE54H30v_bQ1hxI7VMShtELMiIxdAUKnCXcpzwSlQQ"

def get_vehicle_spreadsheet():
    client = get_gsheet_client()
    return client.open_by_key(VEHICLE_SPREADSHEET_ID)

@st.cache_data(ttl=300)
def load_sheet_as_df(sheet_name, columns):
    try:
        sh = get_vehicle_spreadsheet() 
        ws = sh.worksheet(sheet_name)
        data = ws.get_all_records()

        df = pd.DataFrame(data)

        for col in columns:
            if col not in df.columns:
                df[col] = ""

        df = df[columns].copy()

        # None / NaN / nan 문자 방어
        df = df.where(pd.notnull(df), "")
        df = df.replace(["None", "none", "nan", "NaN", "NaT"], "")

        df = df.astype(str)

        return df

    except Exception as e:
        st.error(f"{sheet_name} 데이터를 불러오는 중 오류가 발생했습니다: {e}")
        return pd.DataFrame(columns=columns)


def save_df_to_sheet(sheet_name, df, columns):
    try:
        sh = get_vehicle_spreadsheet()
        ws = sh.worksheet(sheet_name)

        save_df = df.copy()

        for col in columns:
            if col not in save_df.columns:
                save_df[col] = ""

        save_df = save_df[columns].fillna("").astype(str)

        ws.clear()
        ws.update([columns] + save_df.values.tolist())

        return True

    except Exception as e:
        st.error(f"{sheet_name} 저장 중 오류가 발생했습니다: {e}")
        return False


def parse_money(value):
    try:
        if pd.isna(value):
            return 0
        text = str(value).replace(",", "").replace("₩", "").replace("￦", "").strip()
        if text == "":
            return 0
        return int(float(text))
    except:
        return 0


def parse_insurance_end_date(value):
    try:
        text = str(value).strip()
        if not text:
            return None

        if "~" in text:
            text = text.split("~")[-1].strip()

        text = text.replace(".", "-").replace("/", "-")
        dt = pd.to_datetime(text, errors="coerce")

        if pd.isna(dt):
            return None

        return dt.date()

    except:
        return None


def vehicle_page():
    page_title("🚗 차량관리")

    vehicle_df = load_sheet_as_df("차량관리", VEHICLE_COLUMNS)
    repair_df = load_sheet_as_df("차량정비이력", REPAIR_COLUMNS)

    # 숫자 계산용
    vehicle_temp = vehicle_df.copy()
    repair_temp = repair_df.copy()

    vehicle_temp["보험금액_숫자"] = vehicle_temp["보험금액"].apply(parse_money)
    repair_temp["금액_숫자"] = repair_temp["금액"].apply(parse_money)

    total_vehicle = len(vehicle_df)
    active_vehicle = len(vehicle_df[vehicle_df["차량상태"].astype(str).str.strip() != "매각"])
    sold_vehicle = len(vehicle_df[vehicle_df["차량상태"].astype(str).str.strip() == "매각"])
    total_insurance = int(vehicle_temp["보험금액_숫자"].sum())
    total_repair = int(repair_temp["금액_숫자"].sum())

    # 보험 만료 30일 이내 차량 수
    today = date.today()
    insurance_warning_count = 0
    insurance_expired_count = 0

    for _, row in vehicle_df.iterrows():
        end_date = parse_insurance_end_date(row.get("보험기간", ""))

        if end_date:
            remain_days = (end_date - today).days

            if remain_days < 0:
                insurance_expired_count += 1
            elif remain_days <= 30:
                insurance_warning_count += 1

    def vehicle_card(title, value, sub=""):
        ui_card(title, value, sub)

    col1, col2, col3, col4, col5, col6, col7 = st.columns(7)

    with col1:
        vehicle_card("전체 차량", total_vehicle, "등록 차량")

    with col2:
        vehicle_card("운행 차량", active_vehicle, "현재 운행")

    with col3:
        vehicle_card("매각 차량", sold_vehicle, "매각 처리")

    with col4:
        vehicle_card("보험금 합계", f"{total_insurance:,}원", "전체 보험료")

    with col5:
        vehicle_card("정비비 합계", f"{total_repair:,}원", "누적 정비비")

    with col6:
        vehicle_card("보험 만료임박", insurance_warning_count, "30일 이내")

    with col7:
        vehicle_card("보험 만료", insurance_expired_count, "기간 경과")

    if insurance_expired_count > 0:
        st.error(f"🚨 보험이 만료된 차량이 {insurance_expired_count}대 있습니다.")

    if insurance_warning_count > 0:
        st.warning(f"⚠️ 30일 이내 보험 만료 예정 차량이 {insurance_warning_count}대 있습니다.")

    st.divider()

    tab1, tab2, tab3 = st.tabs(["차량 목록", "정비 이력", "보험 만료 경고"])

    # =====================================================
    # 1. 차량 목록
    # =====================================================
    with tab1:
        page_title("차량 목록")

        edited_vehicle_df = st.data_editor(
            vehicle_df,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="vehicle_editor_main_v3",
        )

        if st.button("💾 차량관리 저장", use_container_width=True):
            if save_df_to_sheet(VEHICLE_SHEET_NAME, edited_vehicle_df, VEHICLE_COLUMNS):
                load_sheet_as_df.clear()
                st.success("차량관리 저장 완료!")
                st.rerun()

    # =====================================================
    # 2. 정비 이력
    # =====================================================
    with tab2:
        st.markdown(
            '<div class="erp-section-title">정비 이력</div>',
            unsafe_allow_html=True
        )

        vehicle_numbers = (
            vehicle_df["차량번호"]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .unique()
            .tolist()
        )

        with st.expander("➕ 정비이력 등록", expanded=False):

            with st.form("repair_add_form"):
                r1, r2, r3 = st.columns(3)

                add_vehicle_no = r1.selectbox(
                    "차량번호",
                    vehicle_numbers,
                    key="repair_add_vehicle_no"
                )

                add_repair_date = r2.date_input(
                    "수리일자",
                    value=date.today(),
                    key="repair_add_date"
                )

                add_amount = r3.text_input(
                    "금액",
                    placeholder="예: 120000 또는 무상교체",
                    key="repair_add_amount"
                )

                add_repair_content = st.text_input("수리내역")
                add_note = st.text_input("비고")

                submitted = st.form_submit_button("정비이력 등록")

                # ✅ 반드시 여기 안에!
                if submitted:
                    new_row = pd.DataFrame([{
                        "차량번호": str(add_vehicle_no).strip(),
                        "수리일자": str(add_repair_date),
                        "수리내역": add_repair_content,
                        "금액": add_amount,
                        "비고": add_note
                    }])

                    save_df = pd.concat([repair_df, new_row], ignore_index=True)

                    if save_df_to_sheet(VEHICLE_REPAIR_SHEET_NAME, save_df, REPAIR_COLUMNS):
                        load_sheet_as_df.clear()
                        st.success("정비이력 등록 완료!")
                        st.rerun()

        selected_vehicle = st.selectbox(
            "차량번호 선택",
            ["전체"] + vehicle_numbers
        )

        view_repair_df = repair_df.copy()

        if selected_vehicle != "전체":
            view_repair_df = view_repair_df[
                view_repair_df["차량번호"].astype(str).str.strip() == selected_vehicle
            ]

        with st.expander("📋 정비이력 조회", expanded=False):

            edited_repair_df = st.data_editor(
                view_repair_df,
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key="repair_editor_main_v4"
            )

            if selected_vehicle != "전체":
                st.warning("특정 차량 조회 중에는 전체 정비이력 저장이 비활성화됩니다.")
            else:
                if st.button("💾 정비이력 저장", use_container_width=True):
                    if save_df_to_sheet(VEHICLE_REPAIR_SHEET_NAME, edited_repair_df, REPAIR_COLUMNS):
                        load_sheet_as_df.clear()
                        st.success("정비이력 저장 완료!")
                        st.rerun()

        view_repair_temp = view_repair_df.copy()
        view_repair_temp["금액_숫자"] = view_repair_temp["금액"].apply(parse_money)
        selected_total_repair = int(view_repair_temp["금액_숫자"].sum())

        st.info(f"선택 차량 정비비 합계: {selected_total_repair:,}원")

        repair_edit_df = repair_df.copy()
        repair_edit_df = repair_edit_df.where(pd.notnull(repair_edit_df), "")
        repair_edit_df = repair_edit_df.replace(["None", "none", "nan", "NaN", "NaT"], "")

        for col in REPAIR_COLUMNS:
            if col not in repair_edit_df.columns:
                repair_edit_df[col] = ""

        with st.expander("⚠️ 정비이력 삭제", expanded=False):

            if view_repair_df.empty:
                st.info("삭제할 정비이력이 없습니다.")
            else:
                delete_options = []

                for idx, row in view_repair_df.iterrows():
                    delete_options.append(
                        f"{idx} | {row.get('차량번호', '')} | {row.get('수리일자', '')} | {row.get('수리내역', '')} | {row.get('금액', '')}"
                    )

                selected_delete = st.selectbox(
                    "삭제할 정비이력 선택",
                    delete_options,
                    key="repair_delete_select"
                )

                confirm_delete = st.checkbox(
                    "정말 삭제합니다. 되돌리기 어렵습니다.",
                    key="repair_delete_confirm"
                )

                if st.button("🗑️ 선택 정비이력 삭제", use_container_width=True, key="repair_delete_btn"):
                    if not confirm_delete:
                        st.warning("삭제 확인 체크를 먼저 해주세요.")
                    else:
                        delete_idx = int(selected_delete.split("|")[0].strip())

                        save_df = repair_df[REPAIR_COLUMNS].copy()
                        save_df = save_df.drop(index=delete_idx).reset_index(drop=True)

                        if save_df_to_sheet(VEHICLE_REPAIR_SHEET_NAME, save_df, REPAIR_COLUMNS):
                            load_sheet_as_df.clear()
                            st.success("정비이력이 삭제되었습니다.")
                            st.rerun()

        with st.expander("📊 차량별 정비비 합계", expanded=False):

            if not view_repair_df.empty:

                repair_summary = view_repair_df.copy()
                repair_summary["금액"] = repair_summary["금액"].apply(parse_money)

                summary_df = (
                    repair_summary
                    .groupby("차량번호", as_index=False)["금액"]
                    .sum()
                    .sort_values("금액", ascending=False)
                )

                summary_df["금액"] = summary_df["금액"].apply(lambda x: f"{int(x):,}원")

                st.dataframe(summary_df, use_container_width=True, hide_index=True)

            else:
                st.info("선택한 차량의 정비이력이 없습니다.")


    # =====================================================
    # 3. 보험 만료 경고
    # =====================================================
    with tab3:
        st.markdown(
            '<div class="erp-section-title">보험 만료 경고</div>',
            unsafe_allow_html=True
        )

        today = date.today()
        warning_rows = []

        for _, row in vehicle_df.iterrows():
            end_date = parse_insurance_end_date(row.get("보험기간", ""))

            if end_date:
                remain_days = (end_date - today).days

                if 0 <= remain_days <= 30:
                    warning_rows.append({
                        "차량명": row.get("차량명", ""),
                        "차량번호": row.get("차량번호", ""),
                        "보험회사": row.get("보험회사", ""),
                        "보험기간": row.get("보험기간", ""),
                        "남은일수": remain_days,
                        "비고": row.get("비고", "")
                    })

        if warning_rows:
            st.warning(f"보험 만료 30일 이내 차량이 {len(warning_rows)}대 있습니다.")
            st.dataframe(pd.DataFrame(warning_rows), use_container_width=True, hide_index=True)
        else:
            st.success("보험 만료 임박 차량이 없습니다.")

@st.cache_data(ttl=300)
def load_inspection_data():
    try:
        sheet = get_inspection_sheet()
        values = sheet.get_all_values()

        if not values:
            return pd.DataFrame(columns=INSPECTION_COLUMNS)

        header = values[0]
        data_rows = values[1:]

        # 👉 기존 데이터 그대로 로드
        temp_df = pd.DataFrame(data_rows, columns=header)

        # 👉 새로운 DF 생성
        df = pd.DataFrame(columns=INSPECTION_COLUMNS)

        # 👉 컬럼 매핑 (안전 버전)
        column_map = {
            "요청일": "요청일",
            "운영사": "운영사",
            "현장명": "현장명",
            "단지명": "현장명",
            "현장주소": "현장주소",
            "주소": "현장주소",
            "현장연락처": "현장연락처",
            "전화번호": "현장연락처",
            "주차면수": "주차면수",
            "상품구분": "상품구분",
            "환경부": "환경부",
            "자투": "자투",
            "신규설치수량": "신규설치수량",
            "수량": "신규설치수량",
            "기설치수량": "기설치수량",
            "영업담당자": "영업담당자",
            "영업담당연락처": "영업담당연락처",
            "요청내용": "요청내용",
            "비고": "비고",
            "첨부파일명": "첨부파일명",
            "첨부파일링크": "첨부파일링크",
            "실사담당자": "실사담당자",
            "실사예정일": "실사예정일",
            "실사완료일": "실사완료일",
            "진행상태": "진행상태",
            "실사결과": "실사결과",
            "특이사항": "특이사항",
            "후속조치": "후속조치",
            "계약여부": "계약여부",
            "계약일": "계약일",
            "계약수량": "계약수량",
            "계약금액": "계약금액",
            "미계약사유": "미계약사유",
        }

        # 👉 안전하게 매핑
        for old_col, new_col in column_map.items():
            if old_col in temp_df.columns:
                df[new_col] = temp_df[old_col]

        # 👉 없는 컬럼 채우기
        for col in INSPECTION_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        df = normalize_inspection_df(df)
        return df

    except Exception as e:
        st.error(f"실사 데이터를 불러오지 못했습니다: {e}")
        return pd.DataFrame(columns=INSPECTION_COLUMNS)

def apply_product_filter(df):
    if df is None or df.empty:
        return df

    if "상품구분" not in df.columns:
        return df

    if st.session_state.business == "아이센서":
        return df[df["상품구분"].astype(str).str.strip().isin(["아이센서"])].copy()
    else:
        return df[df["상품구분"].astype(str).str.strip().isin(["전기차충전기", "이전설치"])].copy()        


def save_inspection_data(df, sheet=None):
    if sheet is None:
        sheet = get_inspection_sheet()

    save_df = normalize_inspection_df(df).copy()

    # 임시 컬럼 제거
    if "row_id" in save_df.columns:
        save_df = save_df.drop(columns=["row_id"])

    # 빈 데이터 저장 방지
    if save_df.empty:
        raise Exception("실사 데이터가 비어 있어 저장을 중단했습니다.")

    def clean_cell(x):
        if pd.isna(x):
            return ""
        if isinstance(x, (pd.Timestamp, datetime, date)):
            return x.strftime("%Y-%m-%d")
        text = str(x).strip()
        if text.lower() in ["none", "nan", "nat"]:
            return ""
        return text

    for col in save_df.columns:
        save_df[col] = save_df[col].apply(clean_cell)

    rows = [save_df.columns.tolist()] + save_df.values.tolist()

    if len(rows) <= 1:
        raise Exception("실사 데이터 행이 없어 저장을 중단했습니다.")

    old_values = sheet.get_all_values()
    old_data_rows = max(0, len(old_values) - 1)

    sheet.update("A1", rows, value_input_option="USER_ENTERED")

    new_data_rows = len(save_df)

    if old_data_rows > new_data_rows:
        blank_rows = old_data_rows - new_data_rows
        start_row = new_data_rows + 2
        end_row = old_data_rows + 1

        # 실사 컬럼은 29개라 AC열까지
        clear_range = f"A{start_row}:AC{end_row}"
        empty_values = [[""] * len(INSPECTION_COLUMNS) for _ in range(blank_rows)]

        sheet.update(clear_range, empty_values, value_input_option="USER_ENTERED")

    load_inspection_data.clear()

# =========================================================
# 유지보수 관리 시스템
# =========================================================
MAINTENANCE_SHEET_NAME = "아이센서유지보수"

MAINTENANCE_COLUMNS = [
    "코드번호",
    "단지명",
    "연락처",
    "지역",
    "영업담당자",
    "수량",
    "단가",
    "계약시작일",
    "계약종료일",
    "총계약금액",
    "계약상태",
    "청구주기",
    "비고",
    "첨부파일명",
    "첨부파일링크"
]


def get_maintenance_sheet():
    client = get_gsheet_client()   # 🔥 이거 반드시 필요
    return client.open(MAINTENANCE_SHEET_NAME).sheet1

def get_maintenance_payment_sheet():
    client = get_gsheet_client()   # 🔥 이것도
    return client.open(MAINTENANCE_PAYMENT_SHEET_NAME).sheet1

def ensure_maintenance_sheet_header(sheet):
    values = sheet.get_all_values()

    if not values:
        end_col = chr(64 + len(MAINTENANCE_COLUMNS)) if len(MAINTENANCE_COLUMNS) <= 26 else "O"
        sheet.update(f"A1:{end_col}1", [MAINTENANCE_COLUMNS])
        return

    header = [str(x).strip() for x in values[0]]

    if header != MAINTENANCE_COLUMNS:
        existing = pd.DataFrame(values[1:], columns=header if header else None)

        for col in MAINTENANCE_COLUMNS:
            if col not in existing.columns:
                existing[col] = ""

        existing = existing[MAINTENANCE_COLUMNS]
        save_maintenance_data(existing, sheet=sheet)


def maintenance_safe_int(value, default=0):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return default
    return int(num)


def maintenance_safe_float(value, default=0.0):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return default
    return float(num)


def format_currency(value):
    num = pd.to_numeric(value, errors="coerce")
    if pd.isna(num):
        return "0"
    return f"{int(round(float(num))):,}"

def style_unpaid_amount(val):
    num = pd.to_numeric(val, errors="coerce")
    if pd.isna(num):
        return ""
    if num > 0:
        return "background-color: #f8d7da; color: #842029; font-weight: bold;"
    return ""

def get_contract_expiring_soon(df, within_days=60):
    if df.empty:
        return pd.DataFrame(columns=df.columns)

    temp_df = df.copy()
    temp_df["계약종료일_dt"] = pd.to_datetime(temp_df["계약종료일"], errors="coerce")

    today = pd.Timestamp(date.today())
    limit_day = today + pd.Timedelta(days=within_days)

    temp_df = temp_df[
        (temp_df["계약상태"].astype(str).str.strip() == "진행중") &
        (temp_df["계약종료일_dt"].notna()) &
        (temp_df["계약종료일_dt"] >= today) &
        (temp_df["계약종료일_dt"] <= limit_day)
    ].copy()

    return temp_df.sort_values("계약종료일_dt")


def calculate_total_contract_amount(qty, unit_price):
    return int(maintenance_safe_int(qty, 0) * maintenance_safe_float(unit_price, 0))


@st.cache_data(ttl=300)
def load_maintenance_data():
    sheet = get_maintenance_sheet()
    ensure_maintenance_sheet_header(sheet)

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    if df.empty:
        return pd.DataFrame(columns=MAINTENANCE_COLUMNS)

    for col in MAINTENANCE_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[MAINTENANCE_COLUMNS].copy()

    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    df["단가"] = pd.to_numeric(df["단가"], errors="coerce").fillna(0)
    df["총계약금액"] = pd.to_numeric(df["총계약금액"], errors="coerce").fillna(0)
    df["계약시작일"] = df["계약시작일"].astype(str)
    df["계약종료일"] = df["계약종료일"].astype(str)
    df["계약상태"] = df["계약상태"].astype(str).replace("", "진행중")
    df["청구주기"] = df["청구주기"].astype(str).replace("", "매월")

    return df


def save_maintenance_data(df, sheet=None):
    if sheet is None:
        sheet = get_maintenance_sheet()

    save_df = df.copy()

    if "row_id" in save_df.columns:
        save_df = save_df.drop(columns=["row_id"])

    for col in MAINTENANCE_COLUMNS:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[MAINTENANCE_COLUMNS].fillna("")

    save_df["수량"] = pd.to_numeric(save_df["수량"], errors="coerce").fillna(0).astype(int)
    save_df["단가"] = pd.to_numeric(save_df["단가"], errors="coerce").fillna(0)
    save_df["총계약금액"] = pd.to_numeric(save_df["총계약금액"], errors="coerce").fillna(0)

    rows = [save_df.columns.tolist()] + save_df.astype(str).values.tolist()

    old_values = sheet.get_all_values()
    old_data_rows = max(0, len(old_values) - 1)

    sheet.update("A1", rows)

    new_data_rows = len(save_df)
    total_cols = len(MAINTENANCE_COLUMNS)

    if old_data_rows > new_data_rows:
        blank_rows = old_data_rows - new_data_rows
        start_row = new_data_rows + 2
        end_row = old_data_rows + 1
        end_col_letter = chr(64 + total_cols) if total_cols <= 26 else "O"
        clear_range = f"A{start_row}:{end_col_letter}{end_row}"
        empty_values = [[""] * total_cols for _ in range(blank_rows)]
        sheet.update(clear_range, empty_values)

    load_maintenance_data.clear()

MAINTENANCE_PAYMENT_SHEET_NAME = "아이센서유지보수_수금관리"

MAINTENANCE_PAYMENT_COLUMNS = [
    "코드번호",
    "단지명",
    "기준년월",
    "청구금액",
    "발행여부",
    "발행일",
    "입금여부",
    "입금일",
    "미수금",
    "영업담당자",
    "계약상태",
    "비고"
]


def get_maintenance_payment_sheet():
    client = get_gsheet_client()
    return client.open(MAINTENANCE_PAYMENT_SHEET_NAME).sheet1


def ensure_maintenance_payment_sheet_header(sheet):
    values = sheet.get_all_values()

    if not values:
        sheet.update("A1:L1", [MAINTENANCE_PAYMENT_COLUMNS])
        return

    header = [str(x).strip() for x in values[0]]

    if header != MAINTENANCE_PAYMENT_COLUMNS:
        existing = pd.DataFrame(values[1:], columns=header if header else None)

        for col in MAINTENANCE_PAYMENT_COLUMNS:
            if col not in existing.columns:
                existing[col] = ""

        existing = existing[MAINTENANCE_PAYMENT_COLUMNS]
        save_maintenance_payment_data(existing, sheet=sheet)


@st.cache_data(ttl=300)
def load_maintenance_payment_data():
    sheet = get_maintenance_payment_sheet()
    ensure_maintenance_payment_sheet_header(sheet)

    data = sheet.get_all_records()
    df = pd.DataFrame(data)

    if df.empty:
        return pd.DataFrame(columns=MAINTENANCE_PAYMENT_COLUMNS)

    for col in MAINTENANCE_PAYMENT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[MAINTENANCE_PAYMENT_COLUMNS].copy()

    df["청구금액"] = pd.to_numeric(df["청구금액"], errors="coerce").fillna(0)
    df["미수금"] = pd.to_numeric(df["미수금"], errors="coerce").fillna(0)

    df["발행여부"] = df["발행여부"].astype(str).replace("", "미발행")
    df["입금여부"] = df["입금여부"].astype(str).replace("", "미입금")
    df["기준년월"] = df["기준년월"].astype(str)
    df["발행일"] = df["발행일"].astype(str)
    df["입금일"] = df["입금일"].astype(str)
    df["계약상태"] = df["계약상태"].astype(str)

    return df


def save_maintenance_payment_data(df, sheet=None):
    if sheet is None:
        sheet = get_maintenance_payment_sheet()

    save_df = df.copy()

    if "row_id" in save_df.columns:
        save_df = save_df.drop(columns=["row_id"])

    for col in MAINTENANCE_PAYMENT_COLUMNS:
        if col not in save_df.columns:
            save_df[col] = ""

    save_df = save_df[MAINTENANCE_PAYMENT_COLUMNS].fillna("")

    save_df["청구금액"] = pd.to_numeric(save_df["청구금액"], errors="coerce").fillna(0)
    save_df["미수금"] = pd.to_numeric(save_df["미수금"], errors="coerce").fillna(0)

    rows = [save_df.columns.tolist()] + save_df.astype(str).values.tolist()

    old_values = sheet.get_all_values()
    old_data_rows = max(0, len(old_values) - 1)

    sheet.update("A1", rows)

    new_data_rows = len(save_df)

    if old_data_rows > new_data_rows:
        blank_rows = old_data_rows - new_data_rows
        start_row = new_data_rows + 2
        end_row = old_data_rows + 1
        clear_range = f"A{start_row}:L{end_row}"
        empty_values = [[""] * len(MAINTENANCE_PAYMENT_COLUMNS) for _ in range(blank_rows)]
        sheet.update(clear_range, empty_values)

    load_maintenance_payment_data.clear()

# =========================================================
# 시스템 점검
# =========================================================
def system_check_page():
    st.title("🛠️ 시스템 점검")
    st.caption("읽기 전용 점검 화면입니다. 저장/삭제는 하지 않습니다.")

    if st.session_state.get("role") != "관리자":
        st.warning("관리자만 접근할 수 있습니다.")
        return

    st.markdown(
        '<div class="erp-section-title">1. 기본 상태</div>',
        unsafe_allow_html=True
    )

    try:
        client = get_gsheet_client()
        st.success("✅ 구글 인증 연결 정상")
    except Exception as e:
        st.error(f"❌ 구글 인증 실패: {e}")
        return

    st.divider()

    st.markdown(
        '<div class="erp-section-title">2. 주요 데이터 점검</div>',
        unsafe_allow_html=True
    )

    check_targets = []

    for business_name, config in BUSINESS_CONFIG.items():
        for sheet_name, url in config.get("sheets", {}).items():
            check_targets.append({
                "사업": business_name,
                "시트명": sheet_name,
                "URL": url
            })

    result_rows = []

    for item in check_targets:
        business_name = item["사업"]
        sheet_name = item["시트명"]
        url = item["URL"]

        try:
            df = load_google_sheet_data(business_name, sheet_name, url)

            row_count = len(df)
            col_count = len(df.columns)
            blank_cols = [c for c in df.columns if str(c).strip() == "" or str(c).startswith("빈컬럼")]
            dup_cols = df.columns[df.columns.duplicated()].tolist()

            result_rows.append({
                "사업": business_name,
                "시트명": sheet_name,
                "행수": row_count,
                "컬럼수": col_count,
                "빈컬럼": ", ".join(map(str, blank_cols)) if blank_cols else "",
                "중복컬럼": ", ".join(map(str, dup_cols)) if dup_cols else "",
                "상태": "정상"
            })

        except Exception as e:
            result_rows.append({
                "사업": business_name,
                "시트명": sheet_name,
                "행수": "",
                "컬럼수": "",
                "빈컬럼": "",
                "중복컬럼": "",
                "상태": f"오류: {e}"
            })

    result_df = pd.DataFrame(result_rows)
    st.dataframe(result_df, use_container_width=True, hide_index=True)

    st.divider()

    st.markdown(
        '<div class="erp-section-title">3. 캐시 관리</div>',
        unsafe_allow_html=True
    )

    if st.button("🔄 캐시 초기화", key="system_cache_clear_btn"):
        st.cache_data.clear()
        st.success("캐시 초기화 완료")

# =========================================================
# 페이지들
# =========================================================

def page_import():
    st.title("📥 데이터 가져오기 / 구글시트 연결")
    st.info(f"현재 사업: {st.session_state.business}")

    sheet_urls = get_current_sheet_urls()

    with st.expander("현재 구글시트 링크 상태 보기", expanded=True):
        st.write(sheet_urls)

    st.divider()
    st.markdown(
        '<div class="erp-section-title">구글시트 연결 확인</div>',
        unsafe_allow_html=True
    )
    for sheet_name, url in sheet_urls.items():
        try:
            df = load_google_sheet_data(st.session_state.business, sheet_name, url)
            if df.empty:
                st.warning(f"{sheet_name}: 링크 미입력 또는 데이터 없음")
            else:
                st.success(f"{sheet_name}: 연결 성공 ({len(df)}건)")
        except Exception as e:
            st.error(f"{sheet_name}: 연결 실패 - {e}")

    st.divider()
    st.markdown(
        '<div class="erp-section-title">현재 사업 전체 데이터 백업 다운로드</div>',
        unsafe_allow_html=True
    )
    export_dict = {key: load_df(key) for key in sheet_urls.keys()}
    all_excel = to_excel_bytes(export_dict)
    st.download_button(
        label="전체 데이터 엑셀 백업 다운로드",
        data=all_excel,
        file_name=f"{st.session_state.business}_전체백업_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"download_all_backup_{st.session_state.business}",
    )


def generic_data_page(title, key, filters_config, search_key):
    page_title(title)
    df = load_df(key)
    df = apply_role_filter(df)

    if df.empty:
        st.warning(f"{title} 데이터가 없습니다.")
        return

    cols = st.columns(4)
    filters = {}

    for i, (label, candidates) in enumerate(filters_config):
        if i >= 4:
            break
        col_name = candidates if isinstance(candidates, str) else get_best_column(df, candidates)
        if col_name and col_name in df.columns:
            options = sorted([x for x in df[col_name].astype(str).unique() if x != ""])

            if not is_admin() and label in ["담당자", "영업담당"] and current_user_name() in options:
                cols[i].selectbox(label, [current_user_name()], index=0, key=f"{search_key}_{label}")
                filters[col_name] = current_user_name()
            else:
                value = cols[i].selectbox(label, ["전체"] + options, key=f"{search_key}_{label}")
                filters[col_name] = value

    keyword = st.text_input("검색", placeholder="키워드 검색", key=search_key)
    df2 = filtered_df(df, filters)
    df2 = keyword_filter(df2, keyword)

    st.write(f"조회 건수: {len(df2)}건")
    styled_dataframe(df2)
    download_section(f"{title}_필터결과", df2, title)

def page_tasks():
    page_title("📝 오늘 할 일")
    df = load_tasks_df()

    new_task = st.text_input("할 일 입력")
    if st.button("할 일 추가"):
        if new_task.strip():
            new_row = {
                "등록일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "작성자": current_user_name(),
                "사업": st.session_state.business,
                "할일": new_task.strip(),
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            save_tasks_df(df)

            st.success("추가 완료")
            st.rerun()

    st.write("---")
    view_df = apply_author_filter(df)

    if view_df.empty:
        st.info("등록된 할 일이 없습니다.")
    else:
        st.dataframe(view_df.iloc[::-1].reset_index(drop=True), use_container_width=True, hide_index=True)

        # =========================
        # 오늘할일 삭제 기능
        # =========================
        if not view_df.empty:
            delete_options = [
                f"{idx} | {row['등록일시']} | {row['작성자']} | {row['할일']}"
                for idx, row in view_df.iterrows()
            ]

            selected_delete = st.selectbox(
                "삭제할 할일 선택",
                delete_options,
                key="task_delete_select"
            )

            if st.button("선택한 할일 삭제", key="task_delete_btn"):
                delete_idx = int(selected_delete.split("|")[0].strip())

                df = df.drop(index=delete_idx).reset_index(drop=True)
                save_tasks_df(df)

                st.cache_data.clear()
                st.success("삭제 완료")
                st.rerun()

def page_schedule():
    page_title("📅 일정 관리")

    if st.button("🔄 일정 새로고침"):
        st.cache_data.clear()
        st.rerun()

    df = load_schedule_df()

    title = st.text_input("일정 제목", key="schedule_title")

    # 먼저 기본값 세팅
    if "schedule_date" not in st.session_state:
        st.session_state.schedule_date = date.today()

    # 그 다음 query 반영
    query_params = st.query_params
    if "date" in query_params:
        st.session_state.schedule_date = pd.to_datetime(query_params["date"]).date()
        st.query_params.clear() 

    # 마지막에 input
    selected_date = st.date_input("날짜", key="schedule_date")

    if st.button("일정 추가"):
        if title.strip():
            new_row = {
                "등록일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "작성자": current_user_name(),
                "사업": st.session_state.business,
                "일정명": title.strip(),
                "날짜": selected_date.strftime("%Y-%m-%d"),
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            save_schedule_df(df)
            st.cache_data.clear()
            st.success("일정 추가 완료")
            st.rerun()

    st.write("---")
    view_df = df.copy()
    page_title("📅 월별 일정 달력")

    today = date.today()

    cal_col1, cal_col2 = st.columns(2)

    with cal_col1:
        cal_year = st.number_input(
            "연도",
            min_value=2020,
            max_value=2100,
            value=today.year,
            step=1,
            key="personal_schedule_calendar_year"
        )

    with cal_col2:
        cal_month = st.selectbox(
            "월",
            list(range(1, 13)),
            index=today.month - 1,
            key="personal_schedule_calendar_month"
        )

    calendar_df = view_df.copy()

    if "날짜" in calendar_df.columns:
        calendar_df["날짜"] = pd.to_datetime(calendar_df["날짜"], errors="coerce")
        calendar_df = calendar_df.dropna(subset=["날짜"])

        calendar_df = calendar_df[
            (calendar_df["날짜"].dt.year == int(cal_year)) &
            (calendar_df["날짜"].dt.month == int(cal_month))
        ].copy()

        schedule_map = {}

        for _, row in calendar_df.iterrows():
            day = row["날짜"].day
            author = str(row.get("작성자", "")).strip()
            title = str(row.get("일정명", "")).strip()

            if author:
                title = f"[{author}] {title}"

            if day not in schedule_map:
                schedule_map[day] = []

            if title:
                schedule_map[day].append(title)

        cal = calendar.Calendar(firstweekday=6)
        month_days = cal.monthdayscalendar(int(cal_year), int(cal_month))

        html = f"""
        <style>
        .calendar-table {{
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }}
        .calendar-table th {{
            background: #f1f5f9;
            padding: 10px;
            border: 1px solid #e2e8f0;
            text-align: center;
        }}
        .calendar-table td {{
            height: 110px;
            vertical-align: top;
            border: 1px solid #e2e8f0;
            padding: 8px;
            background: #ffffff;
        }}
        .calendar-day {{
            font-weight: 700;
            margin-bottom: 6px;
        }}
        .calendar-item {{
            font-size: 12px;
            background: #eff6ff;
            border-radius: 6px;
            padding: 4px 6px;
            margin-bottom: 4px;
            color: #1e3a8a;
        }}
        </style>

        <table class="calendar-table">
        <tr>
        <th>일</th><th>월</th><th>화</th><th>수</th><th>목</th><th>금</th><th>토</th>
        </tr>
        """

        for week in month_days:
            html += "<tr>"
            for day in week:
                if day == 0:
                    html += "<td></td>"
                else:
                    items = schedule_map.get(day, [])
                    item_html = ""

                    for item in items:
                        safe_item = str(item).replace("<", "").replace(">", "")
                        item_html += f'<div class="calendar-item">{safe_item}</div>'

                    clicked_date = f"{cal_year}-{cal_month:02d}-{day:02d}"

                    html += f"""
                    <td>
                        <div class="calendar-day">
                            <span style='color:black;'>
                                {day}
                            </span>
                        </div>
                        {item_html}
                    </td>
                    """

            html += "</tr>"

        html += "</table>"

        components.html(html, height=760, scrolling=False)

    else:
        st.info("날짜 컬럼이 없어 달력을 표시할 수 없습니다.")

    with st.expander(f"📋 일정 상세 목록 / 삭제 ({len(view_df)}건)", expanded=False):
        if view_df.empty:
            st.info("등록된 일정이 없습니다.")
        else:
            temp_df = view_df.copy()
            temp_df["날짜정렬"] = pd.to_datetime(temp_df["날짜"], errors="coerce")
            temp_df = temp_df.sort_values(["날짜정렬", "등록일시"], ascending=[True, False]).drop(columns=["날짜정렬"])

            st.dataframe(temp_df, use_container_width=True, hide_index=True)

            # =========================
            # 일정 삭제 기능
            # =========================
            delete_options = [
                f"{idx} | {row['날짜']} | {row['일정명']} | {row['작성자']}"
                for idx, row in view_df.iterrows()
            ]

            selected_delete = st.selectbox(
                "삭제할 일정 선택",
                delete_options,
                key="schedule_delete_select"
            )

            if st.button("선택한 일정 삭제", key="schedule_delete_btn"):
                delete_idx = int(selected_delete.split("|")[0].strip())

                df = df.drop(index=delete_idx).reset_index(drop=True)
                save_schedule_df(df)

                # st.cache_data.clear()
                st.success("삭제 완료")
                st.rerun()

def page_alerts():
    page_title("🚨 영업 알림")

    tax_df = load_tax_alert_df()
    meeting_df = load_meeting_alert_df()
    schedule_df = load_schedule_df()

    view_tax_df = apply_author_filter(tax_df)
    view_meeting_df = apply_author_filter(meeting_df)
    view_schedule_df = apply_author_filter(schedule_df)

    st.markdown(
        '<div class="erp-section-title">1. 알림 요약</div>',
        unsafe_allow_html=True
    )
    c1, c2, c3 = st.columns(3)

    tax_temp = view_tax_df.copy()
    if not tax_temp.empty:
        tax_temp["예정일_dt"] = pd.to_datetime(tax_temp["예정일"], errors="coerce")
        tax_temp["상태표시"] = tax_temp.apply(lambda r: make_alert_status(r["예정일_dt"], r["상태"]), axis=1)
        tax_pending = tax_temp[tax_temp["상태표시"].isin(["지남", "오늘", "긴급", "임박"])]
    else:
        tax_pending = pd.DataFrame()

    meeting_temp = view_meeting_df.copy()
    if not meeting_temp.empty:
        meeting_temp["입대의일자_dt"] = pd.to_datetime(meeting_temp["입대의일자"], errors="coerce")
        meeting_temp["상태표시"] = meeting_temp.apply(lambda r: make_alert_status(r["입대의일자_dt"], r["상태"]), axis=1)
        meeting_pending = meeting_temp[meeting_temp["상태표시"].isin(["지남", "오늘", "긴급", "임박"])]
    else:
        meeting_pending = pd.DataFrame()

    schedule_temp = view_schedule_df.copy()
    if not schedule_temp.empty:
        schedule_temp["날짜_dt"] = pd.to_datetime(schedule_temp["날짜"], errors="coerce")
        schedule_temp["상태표시"] = schedule_temp["날짜_dt"].apply(lambda x: make_alert_status(x, ""))
        schedule_pending = schedule_temp[schedule_temp["상태표시"].isin(["지남", "오늘", "긴급", "임박"])]
    else:
        schedule_pending = pd.DataFrame()

    with c1:
        ui_card("세금계산서 알림", len(tax_pending), "발행 필요")

    with c2:
        ui_card("임대의 알림", len(meeting_pending), "미확인 일정")

    with c3:
        ui_card("일정 알림", len(schedule_pending), "오늘/긴급 일정")

    st.divider()
    st.markdown(
        '<div class="erp-section-title">2. 세금계산서 발행 알림 등록</div>',
        unsafe_allow_html=True
    )
    col1, col2, col3 = st.columns(3)
    tax_site = col1.text_input("단지명", key="tax_site")
    tax_date = col2.date_input("발행 예정일", key="tax_date")
    tax_note = col3.text_input("비고", key="tax_note")

    if st.button("세금계산서 알림 추가"):
        if tax_site.strip():
            new_row = {
                "등록일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "작성자": current_user_name(),
                "사업": st.session_state.business,
                "단지명": tax_site.strip(),
                "예정일": str(tax_date),
                "상태": "예정",
                "비고": tax_note.strip(),
            }
            tax_df = pd.concat([tax_df, pd.DataFrame([new_row])], ignore_index=True)
            save_tax_alert_df(tax_df)
            st.success("세금계산서 알림이 등록되었습니다.")
            st.rerun()

    st.markdown(
        '<div class="erp-section-title">세금계산서 알림 목록</div>',
        unsafe_allow_html=True
    )
    if view_tax_df.empty:
        st.info("등록된 세금계산서 알림이 없습니다.")
    else:
        view_tax = view_tax_df.copy()
        view_tax["예정일_dt"] = pd.to_datetime(view_tax["예정일"], errors="coerce")
        view_tax["D-Day"] = view_tax["예정일_dt"].apply(get_d_day_label)
        view_tax["상태표시"] = view_tax.apply(lambda r: make_alert_status(r["예정일_dt"], r["상태"]), axis=1)
        view_tax = view_tax.sort_values(["예정일_dt", "등록일시"], ascending=[True, False])
        show_alert_table(view_tax[["단지명", "예정일", "D-Day", "상태", "상태표시", "비고", "작성자"]])

    st.divider()
    st.markdown(
        '<div class="erp-section-title">3. 임대의 알림 등록</div>',
        unsafe_allow_html=True
    )
    m1, m2, m3 = st.columns(3)
    meeting_site = m1.text_input("단지명", key="meeting_site")
    meeting_date = m2.date_input("입대의일자", key="meeting_date")
    meeting_note = m3.text_input("비고", key="meeting_note")

    if st.button("입대의 알림 추가"):
        if meeting_site.strip():
            new_row = {
                "등록일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "작성자": current_user_name(),
                "사업": st.session_state.business,
                "단지명": meeting_site.strip(),
                "입대의일자": str(meeting_date),
                "상태": "예정",
                "비고": meeting_note.strip(),
            }
            meeting_df = pd.concat([meeting_df, pd.DataFrame([new_row])], ignore_index=True)
            save_meeting_alert_df(meeting_df)
            st.success("입대의 알림이 등록되었습니다.")
            st.rerun()

    st.markdown(
        '<div class="erp-section-title">임대의 알림 목록</div>',
        unsafe_allow_html=True
    )

    with st.expander(f"📋 상세 보기 ({len(view_meeting_df)}건)", expanded=False):

        if view_meeting_df.empty:
            st.info("등록된 입대의 알림이 없습니다.")
        else:
            view_meeting = view_meeting_df.copy()
            view_meeting["입대의일자_dt"] = pd.to_datetime(view_meeting["입대의일자"], errors="coerce")
            view_meeting["D-Day"] = view_meeting["입대의일자_dt"].apply(get_d_day_label)
            view_meeting["상태표시"] = view_meeting.apply(
                lambda r: make_alert_status(r["입대의일자_dt"], r["상태"]), axis=1
            )

            view_meeting = view_meeting.sort_values(
                ["입대의일자_dt", "등록일시"], ascending=[True, False]
            )

            show_alert_table(
                view_meeting[["단지명", "입대의일자", "D-Day", "상태", "상태표시", "비고", "작성자"]]
            )


def page_admin_tools():
    st.title("🛠 관리자 도구")
    if not is_admin():
        st.error("이 메뉴는 관리자만 사용할 수 있습니다.")
        return

    if st.session_state.business == "아이센서":
        menu = st.radio("관리 작업 선택", ["영업현황 담당자 일괄 확인", "계약단지 담당자 일괄 확인"])

        if menu == "영업현황 담당자 일괄 확인":
            df = load_df("영업현황")
            if df.empty:
                st.info("데이터가 없습니다.")
                return
            site_col = get_site_column(df)
            manager_col = get_manager_column(df)
            if not site_col or not manager_col:
                st.warning("필수 컬럼을 찾지 못했습니다.")
                return
            st.dataframe(df[[site_col, manager_col]], use_container_width=True, height=500)

        elif menu == "계약단지 담당자 일괄 확인":
            df = load_df("계약단지")
            if df.empty:
                st.info("데이터가 없습니다.")
                return
            site_col = get_best_column(df, ["아파트명", "아파트 명", "단지명", "주소"])
            manager_col = get_manager_column(df)
            if not site_col or not manager_col:
                st.warning("필수 컬럼을 찾지 못했습니다.")
                return
            st.dataframe(df[[site_col, manager_col]], use_container_width=True, height=500)

    else:
        st.info("충전기 1차 버전에서는 계약접수현황만 연결되어 있습니다.")
        df = load_df("계약접수현황")
        if not df.empty:
            st.dataframe(df.head(50), use_container_width=True, height=500)

# =========================================================
# 라우터 관리
# =========================================================
ROUTER_COLUMNS = [
    "라우터사용",
    "라우터개통일",
    "라우터명의이전상태",
    "라우터명의이전일",
    "라우터청구대상",
    "라우터월비용",
    "라우터청구시작월",
    "라우터청구종료월",
    "라우터비고",
]

ROUTER_STATUS_OPTIONS = ["", "이전대기", "이전완료", "이전거부", "해지"]
ROUTER_BILLING_OPTIONS = ["", "청구", "미청구"]
ROUTER_USE_OPTIONS = ["", "예", "아니오"]


def clean_router_text(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalize_year_month(value):
    s = clean_router_text(value)
    if s == "":
        return ""

    s = s.replace(".", "-").replace("/", "-").replace(" ", "")
    dt = pd.to_datetime(s, errors="coerce")

    if pd.notna(dt):
        return dt.strftime("%Y-%m")

    # YYYY-MM 형태 직접 허용
    if re.match(r"^\d{4}-\d{1,2}$", s):
        year, month = s.split("-")
        return f"{int(year):04d}-{int(month):02d}"

    return s


def normalize_date_string(value):
    s = clean_router_text(value)
    if s == "":
        return ""
    s = s.replace(".", "-").replace("/", "-").replace(" ", "")
    dt = pd.to_datetime(s, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%Y-%m-%d")
    return s


def router_safe_amount(value):
    s = clean_router_text(value)
    if s == "":
        return 0
    s = s.replace(",", "").replace("원", "").replace(" ", "")
    num = pd.to_numeric(s, errors="coerce")
    if pd.isna(num):
        return 0
    return int(round(float(num)))


def router_yes(value):
    return clean_router_text(value) == "예"


def build_router_base_df():
    if st.session_state.business != "아이센서":
        return pd.DataFrame()

    df = load_df("계약단지")
    df = apply_role_filter(df)

    if df.empty:
        return pd.DataFrame()

    df = df.copy()

    # 라우터 컬럼 없으면 생성
    for col in ROUTER_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # 주요 컬럼 찾기
    site_col = get_best_column(df, ["아파트명", "아파트 명", "단지명", "현장명", "주소"])
    manager_col = get_manager_column(df)
    code_col = get_code_column(df)
    region_col = get_best_column(df, ["지역"])
    qty_col = get_best_column(df, ["수량"])

    if site_col is None:
        site_col = "단지명_표시"
        df[site_col] = ""

    if manager_col is None:
        manager_col = "담당자_표시"
        df[manager_col] = ""

    if code_col is None:
        code_col = "관리코드_표시"
        df[code_col] = ""

    if region_col is None:
        region_col = "지역_표시"
        df[region_col] = ""

    if qty_col is None:
        qty_col = "수량_표시"
        df[qty_col] = ""

    df["단지명_표시"] = df[site_col].astype(str).str.strip()
    df["담당자_표시"] = df[manager_col].astype(str).str.strip()
    df["관리코드_표시"] = df[code_col].astype(str).str.strip()
    df["지역_표시"] = df[region_col].astype(str).str.strip()
    df["수량_표시"] = df[qty_col].astype(str).str.strip()

    # 라우터 컬럼 정규화
    df["라우터사용"] = df["라우터사용"].apply(clean_router_text)
    df["라우터개통일"] = df["라우터개통일"].apply(normalize_date_string)
    df["라우터명의이전상태"] = df["라우터명의이전상태"].apply(clean_router_text)
    df["라우터명의이전일"] = df["라우터명의이전일"].apply(normalize_date_string)
    df["라우터청구대상"] = df["라우터청구대상"].apply(clean_router_text)
    df["라우터월비용"] = df["라우터월비용"].apply(router_safe_amount)
    df["라우터청구시작월"] = df["라우터청구시작월"].apply(normalize_year_month)
    df["라우터청구종료월"] = df["라우터청구종료월"].apply(normalize_year_month)
    df["라우터비고"] = df["라우터비고"].apply(clean_router_text)

    # 날짜형
    df["라우터개통일_dt"] = pd.to_datetime(df["라우터개통일"], errors="coerce")
    df["라우터명의이전일_dt"] = pd.to_datetime(df["라우터명의이전일"], errors="coerce")

    # 표시용 상태 보정
    df["라우터명의이전상태"] = df["라우터명의이전상태"].replace("", "미입력")
    df["라우터청구대상"] = df["라우터청구대상"].replace("", "미입력")
    df["라우터사용"] = df["라우터사용"].replace("", "미입력")

    return df


def get_router_charge_target_df(base_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    if base_df.empty:
        return pd.DataFrame()

    target_ym = f"{int(year):04d}-{int(month):02d}"
    df = base_df.copy()

    # 청구 대상 기본 조건
    df = df[
        (df["라우터사용"] == "예") &
        (df["라우터청구대상"].isin(["청구", "청구대상"])) &
        (df["라우터월비용"] > 0)
    ].copy()

    if df.empty:
        return df

    # 시작월 조건
    start_ok = df["라우터청구시작월"].astype(str).str.strip().eq("") | (df["라우터청구시작월"] <= target_ym)
    df = df[start_ok].copy()

    if df.empty:
        return df

    # 종료월 조건
    end_ok = df["라우터청구종료월"].astype(str).str.strip().eq("") | (df["라우터청구종료월"] >= target_ym)
    df = df[end_ok].copy()

    if df.empty:
        return df

    # 이전완료 제외
    df = df[df["라우터명의이전상태"] != "이전완료"].copy()

    return df


def get_router_issue_df(base_df: pd.DataFrame) -> pd.DataFrame:
    if base_df.empty:
        return pd.DataFrame()

    df = base_df.copy()

    cond = (
        (df["라우터사용"] == "예") &
        (
            ((df["라우터청구대상"].isin(["청구", "청구대상"])) & (df["라우터월비용"] <= 0)) |
            ((df["라우터명의이전상태"] == "이전완료") & (df["라우터청구대상"] == "청구")) |
            ((df["라우터명의이전상태"].isin(["이전대기", "이전거부"])) & (df["라우터청구시작월"] == "")) |
            ((df["라우터사용"] == "예") & (df["라우터청구대상"] == "미입력")) |
            ((df["라우터사용"] == "예") & (df["라우터명의이전상태"] == "미입력"))
        )
    )

    return df[cond].copy()

def get_router_warning_df(base_df: pd.DataFrame, overdue_days: int = 30) -> pd.DataFrame:
    if base_df.empty:
        return pd.DataFrame()

    df = base_df.copy()
    today = pd.Timestamp(datetime.today().date())

    df["개통후경과일"] = None
    if "라우터개통일_dt" in df.columns:
        df["개통후경과일"] = df["라우터개통일_dt"].apply(
            lambda x: (today - x).days if pd.notna(x) else None
        )

    cond_waiting_long = (
        (df["라우터사용"] == "예") &
        (df["라우터명의이전상태"].isin(["이전대기", "이전거부"])) &
        (df["개통후경과일"].notna()) &
        (df["개통후경과일"] >= overdue_days)
    )

    cond_claim_no_amount = (
        (df["라우터사용"] == "예") &
        (df["라우터청구대상"].isin(["청구", "청구대상"])) &
        (pd.to_numeric(df["라우터월비용"], errors="coerce").fillna(0) <= 0)
    )

    cond_done_but_claim = (
        (df["라우터사용"] == "예") &
        (df["라우터명의이전상태"] == "이전완료") &
        (df["라우터청구대상"].isin(["청구", "청구대상"]))
    )

    cond_missing_start = (
        (df["라우터사용"] == "예") &
        (df["라우터청구대상"].isin(["청구", "청구대상"])) &
        (df["라우터청구시작월"].astype(str).str.strip() == "")
    )

    warn_df = df[
        cond_waiting_long | cond_claim_no_amount | cond_done_but_claim | cond_missing_start
    ].copy()

    if warn_df.empty:
        return warn_df

    def make_warning_reason(row):
        reasons = []

        days = row.get("개통후경과일", None)
        status = str(row.get("라우터명의이전상태", "")).strip()
        target = str(row.get("라우터청구대상", "")).strip()
        amount = pd.to_numeric(row.get("라우터월비용", 0), errors="coerce")
        start_ym = str(row.get("라우터청구시작월", "")).strip()

        if status in ["이전대기", "이전거부"] and pd.notna(days) and days >= overdue_days:
            reasons.append(f"개통 후 {int(days)}일 경과")

        if target == "청구" and (pd.isna(amount) or float(amount) <= 0):
            reasons.append("청구대상인데 월비용 0")

        if status == "이전완료" and target == "청구":
            reasons.append("이전완료인데 청구중")

        if target == "청구" and start_ym == "":
            reasons.append("청구시작월 미입력")

        return " / ".join(reasons)

    warn_df["경고사유"] = warn_df.apply(make_warning_reason, axis=1)
    return warn_df


def build_router_claim_export_df(base_df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    target_df = get_router_charge_target_df(base_df, year, month)

    if target_df.empty:
        return pd.DataFrame(columns=[
            "청구년월", "관리코드", "단지명", "담당자", "지역", "수량",
            "라우터명의이전상태", "라우터청구대상", "라우터월비용", "비고"
        ])

    export_df = target_df.copy()
    export_df["청구년월"] = f"{int(year):04d}-{int(month):02d}"

    export_df = export_df[[
        "청구년월",
        "관리코드_표시",
        "단지명_표시",
        "담당자_표시",
        "지역_표시",
        "수량_표시",
        "라우터명의이전상태",
        "라우터청구대상",
        "라우터월비용",
        "라우터비고",
    ]].copy()

    export_df = export_df.rename(columns={
        "관리코드_표시": "관리코드",
        "단지명_표시": "단지명",
        "담당자_표시": "담당자",
        "지역_표시": "지역",
        "수량_표시": "수량",
        "라우터비고": "비고",
    })

    return export_df

def page_router_management():
    page_title("📡 아이센서 라우터 관리")

    if st.session_state.business != "아이센서":
        st.info("라우터 관리는 아이센서 사업에서만 사용합니다.")
        return

    base_df = build_router_base_df()

    if base_df.empty:
        st.warning("계약단지 데이터가 없습니다.")
        return

    current_year = datetime.today().year
    current_month = datetime.today().month
    current_ym = f"{current_year:04d}-{current_month:02d}"

    router_df = base_df[base_df["라우터사용"] == "예"].copy()
    waiting_df = router_df[router_df["라우터명의이전상태"] == "이전대기"].copy()
    rejected_df = router_df[router_df["라우터명의이전상태"] == "이전거부"].copy()
    finished_df = router_df[router_df["라우터명의이전상태"] == "이전완료"].copy()
    charge_target_df = get_router_charge_target_df(base_df, current_year, current_month)
    issue_df = get_router_issue_df(base_df)
    warning_df = get_router_warning_df(base_df, overdue_days=30)

    total_charge_cost = int(charge_target_df["라우터월비용"].sum()) if not charge_target_df.empty else 0

    r1, r2, r3, r4, r5, r6 = st.columns(6)

    with r1:
        ui_card("라우터 사용 단지", len(router_df), "사용중 단지")

    with r2:
        ui_card("이전대기", len(waiting_df), "처리 대기")

    with r3:
        ui_card("이전거부", len(rejected_df), "거부 상태")

    with r4:
        ui_card("이전완료", len(finished_df), "처리 완료")

    with r5:
        ui_card("2026-05 청구대상건수", len(charge_target_df), "청구 대상")

    with r6:
        ui_card("2026-05 청구총액", f"{total_charge_cost:,}", "이번달 청구")
        # 수금 기준 KPI
    billing_df, unpaid_df, manager_df = load_billing_dashboard_data()

    month_billing_df = billing_df[billing_df["기준월"].astype(str).str.strip() == current_ym].copy() if not billing_df.empty else pd.DataFrame()
    month_unpaid_df = unpaid_df[unpaid_df["기준월"].astype(str).str.strip() == current_ym].copy() if not unpaid_df.empty else pd.DataFrame()

    month_unpaid_amount = 0
    month_unpaid_count = 0

    if not month_unpaid_df.empty:
        month_unpaid_amount = int(pd.to_numeric(month_unpaid_df["미수금"], errors="coerce").fillna(0).sum())
        month_unpaid_count = len(month_unpaid_df)

    r7, r8 = st.columns(2)

    with r7:
        ui_card("2026-05 미수금", f"{month_unpaid_amount:,}", "미수금 합계")

    with r8:
        ui_card("2026-05 미입금건수", month_unpaid_count, "미입금 건수")

    if not warning_df.empty:
        st.error(f"경고 항목 {len(warning_df)}건이 있습니다. 아래 '경고/누락 확인'에서 확인하세요.")
    elif not issue_df.empty:
        st.warning(f"입력 보완 필요 항목 {len(issue_df)}건이 있습니다.")
    else:
        st.success("라우터 데이터 상태가 양호합니다.")

    st.divider()

    f1, f2, f3, f4 = st.columns(4)

    manager_options = ["전체"] + sorted([x for x in router_df["담당자_표시"].astype(str).unique().tolist() if x.strip() != ""])
    status_options = ["전체"] + sorted([x for x in router_df["라우터명의이전상태"].astype(str).unique().tolist() if x.strip() != ""])
    billing_options = ["전체"] + sorted([x for x in router_df["라우터청구대상"].astype(str).unique().tolist() if x.strip() != ""])
    region_options = ["전체"] + sorted([x for x in router_df["지역_표시"].astype(str).unique().tolist() if x.strip() != ""])

    sel_manager = f1.selectbox("담당자", manager_options, key="router_manager_filter")
    sel_status = f2.selectbox("명의이전상태", status_options, key="router_status_filter")
    sel_billing = f3.selectbox("청구대상", billing_options, key="router_billing_filter")
    sel_region = f4.selectbox("지역", region_options, key="router_region_filter")

    keyword = st.text_input("단지명 / 관리코드 / 비고 검색", key="router_keyword")

    view_df = router_df.copy()

    if sel_manager != "전체":
        view_df = view_df[view_df["담당자_표시"] == sel_manager]
    if sel_status != "전체":
        view_df = view_df[view_df["라우터명의이전상태"] == sel_status]
    if sel_billing != "전체":
        view_df = view_df[view_df["라우터청구대상"] == sel_billing]
    if sel_region != "전체":
        view_df = view_df[view_df["지역_표시"] == sel_region]

    if keyword.strip():
        kw = keyword.strip()
        view_df = view_df[
            view_df["단지명_표시"].astype(str).str.contains(kw, case=False, na=False) |
            view_df["관리코드_표시"].astype(str).str.contains(kw, case=False, na=False) |
            view_df["라우터비고"].astype(str).str.contains(kw, case=False, na=False)
        ].copy()

    page_title("1. 라우터 전체 현황")
    show_cols = [
        "관리코드_표시", "단지명_표시", "담당자_표시", "지역_표시", "수량_표시",
        "라우터사용", "라우터개통일", "라우터명의이전상태", "라우터명의이전일",
        "라우터청구대상", "라우터월비용", "라우터청구시작월", "라우터청구종료월", "라우터비고"
    ]
    show_df = view_df[show_cols].copy()
    show_df = show_df.rename(columns={
        "관리코드_표시": "관리코드",
        "단지명_표시": "단지명",
        "담당자_표시": "담당자",
        "지역_표시": "지역",
        "수량_표시": "수량",
    })
    styled_dataframe(show_df)

    st.divider()

    st.markdown(
        f'<div class="erp-section-title">2. {current_ym} 청구 대상</div>',
        unsafe_allow_html=True
    )
    if charge_target_df.empty:
        st.info("이번달 청구 대상이 없습니다.")
    else:
        target_show = charge_target_df[[
            "관리코드_표시", "단지명_표시", "담당자_표시", "지역_표시",
            "라우터명의이전상태", "라우터청구대상", "라우터월비용",
            "라우터청구시작월", "라우터청구종료월", "라우터비고"
        ]].copy()
        target_show = target_show.rename(columns={
            "관리코드_표시": "관리코드",
            "단지명_표시": "단지명",
            "담당자_표시": "담당자",
            "지역_표시": "지역",
        })
        st.dataframe(target_show, use_container_width=True, hide_index=True)
        download_section(f"{current_ym}_라우터청구대상", target_show, f"라우터청구대상_{current_ym}")

    st.divider()

    st.markdown(
        '<div class="erp-section-title">3. 청구 엑셀 자동 생성</div>',
        unsafe_allow_html=True
    )
    g1, g2 = st.columns(2)
    claim_year = g1.number_input("청구 연도", min_value=2024, max_value=2100, value=current_year, step=1, key="router_claim_year")
    claim_month = g2.selectbox("청구 월", list(range(1, 13)), index=current_month - 1, key="router_claim_month")

    claim_export_df = build_router_claim_export_df(base_df, int(claim_year), int(claim_month))
    claim_ym = f"{int(claim_year):04d}-{int(claim_month):02d}"

    if claim_export_df.empty:
        st.info(f"{claim_ym} 청구 생성 대상이 없습니다.")
    else:
        st.success(f"{claim_ym} 청구 생성 대상 {len(claim_export_df)}건 / 합계 {int(claim_export_df['라우터월비용'].sum()):,}원")
        st.dataframe(claim_export_df, use_container_width=True, hide_index=True)

        b1, b2 = st.columns([1, 1])

        with b1:
            download_section(
                f"{claim_ym}_라우터청구리스트",
                claim_export_df,
                f"라우터청구리스트_{claim_ym}"
            )

        with b2:
            if st.button(f"{claim_ym} 이번달 청구 생성", key=f"create_billing_{claim_ym}"):
                result = add_monthly_billing_data(claim_export_df)

                if result["added_count"] > 0:
                    st.success(
                        f"수금관리 생성 완료: 신규 {result['added_count']}건 / "
                        f"중복 제외 {result['duplicate_count']}건"
                    )
                else:
                    st.warning(
                        f"신규 생성 없음. 이미 모두 생성된 월입니다. "
                        f"(중복 제외 {result['duplicate_count']}건)"
                    )
                if st.session_state.get("google_update_msg"):
                                    st.info(st.session_state["google_update_msg"])
                                    
                st.rerun()


        st.markdown("### 3-1. 수금관리")
        current_msg = st.session_state.get("google_update_msg", "")
        if current_msg:
            st.error(current_msg)
            st.session_state["google_update_msg"] = ""

        billing_df, unpaid_df, manager_df = load_billing_dashboard_data()

        # 기존 카드/요약
        if billing_df.empty:
            st.info("수금관리 데이터가 없습니다.")
        else:
            total_unpaid = int(pd.to_numeric(billing_df["미수금"], errors="coerce").fillna(0).sum())
            unpaid_count = int((pd.to_numeric(billing_df["미수금"], errors="coerce").fillna(0) > 0).sum())

            s1, s2 = st.columns(2)

            with s1:
                ui_card("총 미수금", f"{total_unpaid:,}", "전체 미수금")

            with s2:
                ui_card("미입금건수", unpaid_count, "미입금 건수")

            month_billing_df = billing_df[
                billing_df["기준월"].astype(str).str.strip() == claim_ym
            ].copy()

            month_billing_df["입금여부"] = month_billing_df["입금여부"].replace("", "미입금")

            if month_billing_df.empty:
                st.info(f"{claim_ym} 수금관리 데이터가 없습니다.")
            else:
                st.dataframe(month_billing_df, use_container_width=True, hide_index=True)

                st.markdown("### 입금 처리")

                unpaid_candidates = month_billing_df[
                    pd.to_numeric(month_billing_df["미수금"], errors="coerce").fillna(0) > 0
                ].copy()

                if unpaid_candidates.empty:
                    st.success("이 달의 청구 건은 모두 입금 처리되었습니다.")
                else:
                    unpaid_candidates = month_unpaid_df.copy()

                    unpaid_candidates["청구금액"] = pd.to_numeric(
                        unpaid_candidates["청구금액"], errors="coerce"
                    ).fillna(0)

                    unpaid_candidates["미수금"] = pd.to_numeric(
                        unpaid_candidates["미수금"], errors="coerce"
                    ).fillna(unpaid_candidates["청구금액"]).fillna(0)

                    unpaid_candidates["표시금액"] = unpaid_candidates["미수금"]
                    unpaid_candidates.loc[unpaid_candidates["표시금액"] <= 0, "표시금액"] = unpaid_candidates["청구금액"]

                    unpaid_candidates["표시금액"] = unpaid_candidates["표시금액"].astype(int)

                    unpaid_candidates["선택표시"] = (
                        unpaid_candidates["단지명"].astype(str).str.strip()
                        + " / "
                        + unpaid_candidates["담당자"].astype(str).str.strip()
                        + " / "
                        + unpaid_candidates["표시금액"].astype(str).str.strip()
                        + "원"
                    )

                    selected_label = st.selectbox(
                        "입금 처리할 항목 선택",
                        unpaid_candidates["선택표시"].tolist(),
                        key=f"paid_select_{claim_ym}"
                    )

                    selected_row = unpaid_candidates[unpaid_candidates["선택표시"] == selected_label].iloc[0]

                    if st.button("선택 항목 입금 처리", key=f"mark_paid_{claim_ym}"):
                        st.session_state["google_update_msg"] = ""

                        amount_raw = selected_row.get("표시금액", "")
                        amount_num = pd.to_numeric(amount_raw, errors="coerce")

                        if pd.isna(amount_num):
                            st.error(f"청구금액이 비어 있거나 숫자가 아닙니다: {amount_raw}")
                        else:
                            ok = mark_billing_paid(
                                기준월=str(selected_row.get("기준월", "")).strip(),
                                단지명=str(selected_row.get("단지명", "")).strip(),
                                담당자=str(selected_row.get("담당자", "")).strip(),
                                청구금액=int(amount_num)
                            )

                            if ok:
                                st.success("입금 처리 완료")
                                st.rerun()
                            else:
                                st.error("입금 처리 실패")  

        st.markdown("### 3-2. 미입금관리")
        if unpaid_df.empty:
            st.success("현재 미입금 항목이 없습니다.")
        else:
            month_unpaid_df = unpaid_df[unpaid_df["기준월"].astype(str).str.strip() == claim_ym].copy()
            st.dataframe(month_unpaid_df, use_container_width=True, hide_index=True)

        st.markdown("### 3-3. 담당자별현황")
        if manager_df.empty:
            st.info("담당자별현황 데이터가 없습니다.")
        else:
            st.dataframe(manager_df, use_container_width=True, hide_index=True)

    st.divider()

    st.markdown(
        '<div class="erp-section-title">6. 경고 / 누락 확인</div>',
        unsafe_allow_html=True
    )
    if warning_df.empty:
        st.success("긴급 경고 항목이 없습니다.")
    else:
        warn_show = warning_df[[
            "관리코드_표시", "단지명_표시", "담당자_표시", "지역_표시",
            "라우터명의이전상태", "라우터청구대상", "라우터월비용",
            "라우터청구시작월", "라우터청구종료월", "경고사유"
        ]].copy()
        warn_show = warn_show.rename(columns={
            "관리코드_표시": "관리코드",
            "단지명_표시": "단지명",
            "담당자_표시": "담당자",
            "지역_표시": "지역",
        })
        st.dataframe(warn_show, use_container_width=True, hide_index=True)
        download_section("라우터_경고항목", warn_show, "라우터_경고항목")

    st.divider()
    st.info("※ 현재 화면은 계약단지 구글시트의 라우터 컬럼을 읽어서 보여주는 관리 화면입니다. 값 수정은 원본 구글시트에서 진행해 주세요.")

def is_admin():
    role = str(st.session_state.get("role", "")).strip()
    return role in ["관리자", "admin", "ADMIN"]


def is_admin():
    role = str(st.session_state.get("role", "")).strip()
    return role in ["관리자", "admin", "ADMIN"]

def current_user_name():
    return st.session_state.display_name or st.session_state.username

def logout():
    st.session_state.logged_in = False
    st.session_state.username = ""
    st.session_state.role = ""
    st.session_state.display_name = ""
    st.session_state.department = ""
    st.session_state.position = ""
    st.session_state.user_code = ""
    st.rerun()

# =========================================================
# 메인
# =========================================================
def main():
    render_common_style()   
    render_header() 
    init_files()

    st.sidebar.title("메뉴")
    st.sidebar.write(f"로그인: {st.session_state.username}")
    st.sidebar.write(f"이름: {current_user_name()}")
    st.sidebar.write(f"권한: {'관리자' if is_admin() else '담당자'}")

    if st.session_state.get("department"):
        st.sidebar.write(f"부서: {st.session_state.department}")

    if st.session_state.get("position"):
        st.sidebar.write(f"직급: {st.session_state.position}")

    if st.session_state.get("user_code"):
        st.sidebar.write(f"코드: {st.session_state.user_code}")

    selected_business = st.sidebar.selectbox("사업 선택", list(BUSINESS_CONFIG.keys()))
    st.session_state.business = selected_business

    if st.sidebar.button("로그아웃"):
        logout()

    if st.sidebar.button("🏠 홈 ", use_container_width=True):
        st.session_state["sidebar_menu_group"] = "📊 통합"
        st.session_state["sidebar_menu_item"] = "대시보드"
        st.rerun()

    # =====================================================
    # 사이드바 메뉴 그룹화
    # =====================================================
    if st.session_state.business == "아이센서":

        menu_groups = {
            "📊 통합": ["대시보드", "시스템 점검"],
            "📁 영업관리": [
                "영업현황",
                "가능단지",
                "입찰공고단지",
                "계약단지",
            ],
            "🛠 시공/실사관리": [
                "실사 관리",
                "시공 일정",
            ],
            "👥 인사관리": [
                "연차 관리",
            ],
            "✅ 업무관리": [
                "오늘 할 일",
                "일정 관리",
                "영업 알림",
            ],
        }

        # ✅ 관리자만 유지보수 추가
        if is_admin():
            menu_groups["🔬 유지보수관리"] = [
                "라우터 관리",
                "아이센서 유지보수관리",
                "차량관리"
            ]

        if is_admin():
            menu_groups["⚙ 관리자"] = [
                "데이터 가져오기",
                "관리자 도구",
            ]

    else:
        menu_groups = {
            "📊 통합": [
                "대시보드",
            ],
            "📁 영업관리": [
                "계약접수현황",
            ],
            "🛠 시공/실사관리": [
                "실사 관리",
                "시공 일정",
            ],
            "✅ 업무관리": [
                "오늘 할 일",
                "일정 관리",
                "영업 알림",
            ],
        }

        if is_admin():
            menu_groups["⚙ 관리자"] = [
                "데이터 가져오기",
                "관리자 도구",
            ]

    st.sidebar.markdown("### 업무 구분")

    selected_group = st.sidebar.selectbox(
        "카테고리",
        list(menu_groups.keys()),
        key="sidebar_menu_group"
    )

    menu = st.sidebar.radio(
        "메뉴 선택",
        menu_groups[selected_group],
        key="sidebar_menu_item"
    )

    if menu == "대시보드":
        page_dashboard()

    elif menu == "시스템 점검":
        system_check_page()    

    elif menu == "데이터 가져오기":
        page_import()

    elif menu == "영업현황":
        generic_data_page(
            "📞 영업현황",
            "영업현황",
            [("담당자", ["담당자", "영업담당"]), ("지역", "지역"), ("상품", "상품"), ("진행여부", "진행여부")],
            "sales_search",
        )

    elif menu == "가능단지":
        generic_data_page(
            "✅ 가능단지",
            "가능단지",
            [("영업담당", ["담당자", "영업담당"]), ("지역", "지역"), ("상품", "상품"), ("결과", "결과")],
            "possible_search",
        )

    elif menu == "입찰공고단지":
        generic_data_page(
            "📝 입찰공고단지",
            "입찰공고",
            [("영업담당", ["담당자", "영업담당"]), ("지역", "지역"), ("상품", "상품"), ("판매형태", "판매형태")],
            "bid_search",
        )

    elif menu == "계약단지":
        generic_data_page(
            "📦 계약단지",
            "계약단지",
            [("영업담당", ["담당자", "영업담당"]), ("지역", "지역"), ("상품", "상품"), ("시공여부", "시공여부")],
            "contract_search",
        )

    elif menu == "라우터 관리":
        page_router_management()    

    elif menu == "연차 관리":
        vacation_page()

    elif menu == "시공 일정":
        schedule_page()

    elif menu == "실사 관리":
        inspection_page()

    elif menu == "아이센서 유지보수관리":
        maintenance_page()

    elif menu == "차량관리":
        vehicle_page()            

    elif menu == "계약접수현황":
        generic_data_page(
            "🔋 충전기 계약접수현황",
            "계약접수현황",
            [
                ("담당자", ["담당자", "영업담당", "실사담당"]),
                ("지역", "지역"),
                ("구분", "구분"),
                ("운영사", "운영사"),
            ],
            "ev_contract_search",
        )

    elif menu == "오늘 할 일":
        page_tasks()

    elif menu == "일정 관리":
        page_schedule()

    elif menu == "영업 알림":
        page_alerts()

    elif menu == "관리자 도구":
        page_admin_tools()


if not st.session_state.logged_in:
    login()
else:
    main()
